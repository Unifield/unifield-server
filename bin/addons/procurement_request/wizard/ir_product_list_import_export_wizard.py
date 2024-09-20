#!/usr/bin/env python
#-*- encoding:utf-8 -*-

from osv import osv
from osv import fields
from tools.translate import _
from openpyxl import load_workbook
from io import BytesIO
import base64
import time


class ir_product_list_export_wizard(osv.osv_memory):
    _name = 'ir.product.list.export.wizard'

    _columns = {
        'product_list_id': fields.many2one('product.list', string='Product List'),
    }

    def create_report(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        return {'type': 'ir.actions.report.xml', 'report_name': 'report_ir_product_list_export', 'datas': {'target_filename': _('Product List IR Excel Template')}, 'context': context}


ir_product_list_export_wizard()


class ir_product_list_import_wizard(osv.osv_memory):
    _name = 'ir.product.list.import.wizard'
    _rec_name = 'sale_id'

    _columns = {
        'sale_id': fields.many2one('sale.order', string='Internal Request', required=True, readonly=True),
        'state': fields.selection([('draft', 'Draft'), ('in_progress', 'Import in progress'), ('error', 'Error'), ('done', 'Done')], string='State', readonly=True),
        'file_to_import': fields.binary(string='File to import', filters='*.xls*'),
        'filename': fields.char(size=64, string='Filename'),
        'message': fields.text(string='Message', readonly=True, translate=True),
    }

    _defaults = {
        'state': 'draft'
    }

    def go_to_ir(self, cr, uid, ids, context=None):
        '''
        Return to the initial view.
        '''
        if isinstance(ids, int):
            ids = [ids]
        act_obj = self.pool.get('ir.actions.act_window')
        for wiz in self.browse(cr, uid, ids, context=context):
            res = act_obj.open_view_from_xmlid(cr, uid, 'procurement_request.action_procurement_request', ['form', 'tree'], context=context)
            res['res_id'] = wiz.sale_id.id
            return res

        return act_obj.open_view_from_xmlid(cr, uid, 'procurement_request.action_procurement_request', ['tree', 'form'], context=context)

    def import_ir(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        prod_obj = self.pool.get('product.product')

        start_time = time.time()
        wiz = self.browse(cr, uid, ids[0], context=context)
        sale = wiz.sale_id
        if not wiz.file_to_import:
            raise osv.except_osv(_('Error'), _('Nothing to import.'))
        self.write(cr, uid, wiz.id, {'state': 'in_progress'}, context=context)

        wb = load_workbook(filename=BytesIO(base64.b64decode(wiz.file_to_import)), read_only=True)
        sheet = wb.active

        lines = []
        lines_err = []
        lines_warn = []
        message = ''
        def_line = {'order_id': sale.id}
        for cell in sheet.iter_rows(min_row=10, min_col=1, max_col=4):
            if not cell[0].value:  # Stop looking at lines if there is no product
                break
            line = def_line.copy()
            line_err = ''
            line_warn = ''
            line_num = cell[0].row

            # Product Code
            prod_name = cell[0].value
            prod_ids = prod_obj.search(cr, uid, [('default_code', '=ilike', prod_name)], context=context)
            if not prod_ids:
                lines.append([{}, _('Line %s: There is no active product %s. ') % (line_num, prod_name), ''])
                continue

            # Check constraints on products
            p_error, p_msg = prod_obj._test_restriction_error(cr, uid, [prod_ids[0]],  vals={'constraints': ['consumption']}, context=context)
            if p_error:
                lines.append([{}, _('Line %s: %s. ') % (line_num, p_msg), ''])
                continue

            ftf = ['name', 'standard_price', 'uom_id', 'procure_method']
            prod = prod_obj.browse(cr, uid, prod_ids[0], fields_to_fetch=ftf, context=context)
            line.update({
                'product_id': prod.id,
                'name': prod.name,
                'price_unit': prod.standard_price,
                'uom_id': prod.uom_id.id,
            })

            # Quantity
            qty = cell[2].value
            if qty:
                if cell[2].data_type == 'n':
                    line.update({'product_uom_qty': qty, 'product_uos_qty': qty})
                else:
                    try:
                        qty = float(qty.rstrip().replace(',', '.'))
                        line.update({'product_uom_qty': qty, 'product_uos_qty': qty})
                    except ValueError:
                        line_err += _('The Quantity must be a number. ')
            else:
                line_err += _('The Quantity is mandatory for each line. ')

            # Date of Stock Take
            ds_date = cell[3].value
            if ds_date:
                if ds_date and (cell[3].data_type != 'd' or not cell[3].is_date):
                    line_warn += _('%s: The Date of Stock Take must be a date. ') % (prod_name,)
                    ds_date = False
                if ds_date:
                    try:
                        ds_date = ds_date.strftime('%Y-%m-%d')  # Fix format
                        if ds_date > sale.date_order:
                            line_warn += _('The Date of Stock Take is not consistent! It should not be later than %s\'s creation date. ') % (sale.name,)
                        else:
                            line.update({'stock_take_date': ds_date})
                    except ValueError:
                        line_warn += _('%s the Date of Stock Take %s is not correct. ') % (prod_name, ds_date)

            if line_err:
                line_err = _('Line %s: ') % (line_num, ) + line_err
            if line_warn:
                line_warn = _('Line %s: ') % (line_num, ) + line_warn
            lines.append([line, line_err, line_warn])

        wiz_state = 'done'
        imp_lines = 0
        for line, line_err, line_warn in lines:
            if not line_err:
                self.pool.get('sale.order.line').create(cr, uid, line, context=context)
                imp_lines += 1
            elif line_err:
                lines_err.append(line_err)
            if line_warn:
                lines_warn.append(line_warn)
        if lines_err:
            message = '%s:\n%s' % (_('Errors'), "\n".join(lines_err))
            wiz_state = 'error'

        end_time = time.time()
        total_time = str(round(end_time - start_time)) + _(' second(s)')
        final_message = _(''' 
Importation completed in %s!
# of imported lines : %s on %s lines
# of ignored lines: %s
# of errors to correct: %s

%s''') % (total_time, imp_lines, len(lines), len(lines) - imp_lines, len(lines_err), message)
        if lines_warn:
            final_message += "\n%s:\n%s" % (_('Warning'), "\n".join(lines_warn))

        self.write(cr, uid, wiz.id, {'state': wiz_state, 'message': final_message}, context=context)

        wb.close()  # Close manually because of readonly
        return True


ir_product_list_import_wizard()
