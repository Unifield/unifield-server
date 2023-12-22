# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2010 Tiny SPRL (<http://tiny.be>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

import time
import threading
import base64

from osv import osv
from osv import fields
from spreadsheet_xml.xlsx_write import XlsxReport
from spreadsheet_xml.xlsx_write import XlsxReportParser
from tools.translate import _
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from service.web_services import report_spool

from datetime import datetime
from dateutil.relativedelta import relativedelta


IN_LAST_X_MONTHS = tuple([("%s" % i, _("%s months") % str(i)) for i in range(1, 13)])


class export_report_stock_inventory(osv.osv):
    _name = 'export.report.stock.inventory'
    _order = 'id desc'

    _columns = {
        'company_id': fields.many2one('res.company', string='DB/Instance name', readonly=True),
        'name': fields.datetime(string='Report Generation date', readonly=True),
        'stock_level_date': fields.date(string='Stock Level date'),
        'product_id': fields.many2one('product.product', string='Specific product', help="If a product is chosen, only quantities of this product will be shown."),
        'prodlot_id': fields.many2one('stock.production.lot', string='Specific Batch number'),
        'product_list_id': fields.many2one('product.list', string='Specific Product list'),
        'nomen_family_id': fields.many2one('product.nomenclature', 'Specific Product family', select=1, domain=[('level', '=', 2)]),
        'mml_id': fields.many2one('msf.instance', string='MML'),
        'msl_id': fields.many2one('unifield.instance', domain=[('uf_active', '=', True)], string='MSL'),
        'expiry_date': fields.date(string='Specific expiry date'),
        'location_id': fields.many2one('stock.location', string='Specific location', help="If a location is chosen, only product quantities in this location will be shown.", required=False),
        'state': fields.selection(selection=[('draft', 'Draft'), ('in_progress', 'In Progress'), ('ready', 'Ready')], string='State', readonly=True,),
        'display_0': fields.boolean(string='Include products with stock <= 0 with movements in the last months'),
        'in_last_x_months': fields.selection(IN_LAST_X_MONTHS, 'In the last'),
    }

    _defaults = {
        'state': lambda *a: 'draft',
        'company_id': lambda s, cr, uid, c: s.pool.get('res.company')._company_default_get(cr, uid, 'export.report.stock.inventory', context=c),
    }

    def onchange_sl_date(self, cr, uid, ids, sl_date=False):
        if sl_date and sl_date > str(datetime.now().date()):
            return {'value': {'stock_level_date': False}}
        else:
            return {'value': {'stock_level_date': sl_date}}

    def update(self, cr, uid, ids, context=None):
        return {}

    def generate_report_all_locations(self, cr, uid, ids, context=None):
        """
        Launch the generation of a report with the inventory of each products on each locations
        @param cr: Cursor to the database
        @param uid: ID of the res.users that calls this method
        @param ids: List of ID of export wizard
        @param context: Context of the call
        @return: An action to the view with a waiting message
        """
        return self.generate_report(cr, uid, ids, context=context, all_locations=True)

    def generate_report(self, cr, uid, ids, context=None, all_locations=False):
        """
        Select the good lines on the report.stock.inventory table
        """
        data_obj = self.pool.get('ir.model.data')

        if context is None:
            context = {}

        for report in self.browse(cr, uid, ids, context=context):
            self.write(cr, uid, [report.id], {
                'name': time.strftime('%Y-%m-%d %H:%M:%S'),
                'state': 'in_progress',
            })

            cr.commit()
            new_thread = threading.Thread(
                target=self.generate_report_bkg,
                args=(cr, uid, report.id, context, all_locations)
            )
            new_thread.start()
            new_thread.join(30.0)

            res = {
                'type': 'ir.actions.act_window',
                'res_model': self._name,
                'view_type': 'form',
                'view_mode': 'form,tree',
                'res_id': report.id,
                'context': context,
                'target': 'same',
            }
            if new_thread.is_alive():
                view_id = data_obj.get_object_reference(
                    cr, uid,
                    'specific_rules',
                    'export_report_stock_inventory_info_view')[1]
                res['view_id'] = [view_id]

            return res

        raise osv.except_osv(
            _('Error'),
            _('No inventory lines found for these parameters'),
        )

    def generate_report_bkg(self, cr, uid, ids, context=None, all_locations=False):
        """
        Generate the report in background
        """
        if context is None:
            context = {}

        if isinstance(ids, int):
            ids = [ids]

        import pooler
        new_cr = pooler.get_db(cr.dbname).cursor()

        report_name = 'report_stock_inventory_xlsx'
        attachment_name = 'stock_inventory_location_view_%s.xlsx'
        if all_locations:
            attachment_name = 'stock_inventory_global_view_%s.xlsx'

        rp_spool = report_spool()
        result = rp_spool.exp_report(cr.dbname, uid, report_name, ids, {'report_id': ids[0]}, context)

        file_res = {'state': False}
        while not file_res.get('state'):
            file_res = rp_spool.exp_report_get(new_cr.dbname, uid, result)
            time.sleep(0.5)
        attachment = self.pool.get('ir.attachment')
        attachment.create(new_cr, uid, {
            'name': attachment_name % (
                time.strftime('%Y_%m_%d_%H_%M')),
            'datas_fname': attachment_name % (
                time.strftime('%Y_%m_%d_%H_%M')),
            'description': 'Stock inventory',
            'res_model': 'export.report.stock.inventory',
            'res_id': ids[0],
            'datas': base64.b64encode(open(file_res.get('path'), 'rb').read()),
        })
        self.write(new_cr, uid, ids, {'state': 'ready'}, context=context)

        new_cr.commit()
        new_cr.close(True)

        return True

    def onchange_product_list(self, cr, uid, ids, product_list_id):
        """
        Change the product when change the prodlot
        """
        if product_list_id:
            return {
                'value': {
                    'display_0': False,
                    'in_last_x_months': False,
                },
            }
        return {}

    def onchange_prodlot(self, cr, uid, ids, prodlot_id):
        """
        Change the product when change the prodlot
        """
        if not prodlot_id:
            return {
                'value': {
                    'product_id': False,
                    'expiry_date': False,
                },
            }

        prodlot = self.pool.get('stock.production.lot').\
            browse(cr, uid, prodlot_id)
        return {
            'value': {
                'product_id': prodlot.product_id.id,
                'expiry_date': prodlot.life_date,
            },
        }

    def create(self, cr, uid, vals, context=None):
        """
        Call onchange_prodlot if a lot is defined
        """
        if vals.get('prodlot_id'):
            vals.update(
                self.onchange_prodlot(
                    cr, uid, False, vals.get('prodlot_id')
                ).get('value', {})
            )

        return super(export_report_stock_inventory, self).\
            create(cr, uid, vals, context=context)

    def write(self, cr, uid, ids, vals, context=None):
        """
        Call onchange_prodlot if a lot is defined
        """
        if not ids:
            return True
        if vals.get('prodlot_id'):
            vals.update(
                self.onchange_prodlot(
                    cr, uid, ids, vals.get('prodlot_id')
                ).get('value', {})
            )

        return super(export_report_stock_inventory, self).\
            write(cr, uid, ids, vals, context=context)


export_report_stock_inventory()


class export_report_stock_inventory_parser(XlsxReportParser):
    def add_cell(self, value=None, style='default_style'):
        # None value set an xls empty cell
        # False value set the xls False()
        new_cell = WriteOnlyCell(self.workbook.active, value=value)
        new_cell.style = style
        self.rows.append(new_cell)

    def get_locations(self, context=None):
        if context is None:
            context = {}

        loc_obj = self.pool.get('stock.location')
        location_ids = loc_obj.search(self.cr, self.uid, [('usage', '=', 'internal')], context=context)
        return loc_obj.browse(self.cr, self.uid, location_ids, context=context)

    def get_lines(self, report_id, context=None):
        if context is None:
            context = {}

        prod_obj = self.pool.get('product.product')

        res = {}
        report = self.pool.get('export.report.stock.inventory').browse(self.cr, self.uid, report_id, context=context)

        with_zero = False
        values = {'state': 'done'}
        cond = ['state=%(state)s']
        having = ['having round(sum(product_qty), 6) != 0']

        full_prod_list = []
        date_prod_list = []  # List of products with moves within the date range of months selected
        batch_list = []

        cond.append('location_id in %(location_ids)s')
        if report.location_id:
            values['location_ids'] = (report.location_id.id,)
        else:
            values['location_ids'] = tuple(self.pool.get('stock.location').search(self.cr, self.uid, [('usage', '=', 'internal')]))

        if report.product_id:
            cond.append('product_id in %(product_ids)s')
            full_prod_list = [report.product_id.id]
            values['product_ids'] = (report.product_id.id,)
        else:
            plist_prod_list = []
            mmlmsl_prod_list = []
            if report.product_list_id:
                plist_prod_list = prod_obj.search(self.cr, self.uid, [('list_ids', '=', report.product_list_id.id)], context=context)
            if report.mml_id or report.msl_id:
                dom = []
                if report.mml_id:
                    dom.append(('in_mml_instance', '=', report.mml_id.id))
                if report.msl_id:
                    dom.append(('in_msl_instance', '=', report.msl_id.id))
                mmlmsl_prod_list = prod_obj.search(self.cr, self.uid, dom, context=context)

            if plist_prod_list and mmlmsl_prod_list:
                full_prod_list = plist_prod_list.intersection(mmlmsl_prod_list)
            elif plist_prod_list and not mmlmsl_prod_list:
                full_prod_list = plist_prod_list
            elif not plist_prod_list and mmlmsl_prod_list:
                full_prod_list = mmlmsl_prod_list
            if full_prod_list:
                cond.append('product_id in %(product_ids)s')
                values['product_ids'] = tuple(full_prod_list)
                with_zero = True

        if report.nomen_family_id:
            cond.append('nomen_family_id=%(nomen_family_id)s')
            values['nomen_family_id'] = report.nomen_family_id.id
            with_zero = True

        if report.prodlot_id:
            cond.append('prodlot_id=%(prodlot_id)s')
            values['prodlot_id'] = report.prodlot_id.id
            with_zero = True

        if report.expiry_date:
            cond.append('expired_date=%(expiry_date)s')
            values['expiry_date'] = report.expiry_date
            with_zero = True

        if report.stock_level_date:
            cond.append('date<%(stock_level_date)s')
            values['stock_level_date'] = '%s 23:59:59' % report.stock_level_date

        if (not report.product_list_id or not report.prodlot_id or not report.expiry_date) and report.display_0:
            to_date = datetime.now()
            if report.stock_level_date:
                to_date = datetime.strptime(values['stock_level_date'], '%Y-%m-%d %H:%M:%S')
            from_date = (to_date + relativedelta(months=-int(report.in_last_x_months))).strftime('%Y-%m-%d 00:00:00')
            with_zero = True

            w_prod = ""
            if report.product_id:
                w_prod = " product_id = %s AND" % report.product_id.id

            self.cr.execute("""
                SELECT DISTINCT m.product_id, m.prodlot_id FROM stock_move m
                LEFT JOIN product_product p ON m.product_id = p.id 
                WHERE""" + w_prod + """ m.state = 'done' AND m.product_qty != 0 AND p.active = 't' AND
                    (location_id IN %s OR location_dest_id IN %s) AND m.date >= %s AND m.date <= %s
                """, (values['location_ids'], values['location_ids'], from_date, to_date))
            for x in self.cr.fetchall():
                full_prod_list.append(x[0])
                date_prod_list.append(x[0])
                if x[1]:
                    batch_list.append(x[1])

        if report.display_0:
            if batch_list:
                having.append('or prodlot_id in %(batch_list)s')
                values['batch_list'] = tuple(batch_list)
            else:
                having.append('or prodlot_id is NULL')
        elif with_zero:
            having = []

        self.cr.execute("""select sum(product_qty), product_id, expired_date, prodlot_id, location_id
            from report_stock_inventory
            where
                """ + ' and '.join(cond) + """ 
            group by product_id, expired_date, uom_id, prodlot_id, location_id
            """ + ' '.join(having), values)

        all_product_ids = {}
        all_bn_ids = {}
        bn_data = {}
        product_data = {}
        # fetch data with db id: for uom, product, bn ...
        for line in self.cr.fetchall():
            if report.display_0 and line[1] not in date_prod_list and line[0] == 0:
                continue
            all_product_ids[line[1]] = True
            all_bn_ids[line[3]] = True

            res.setdefault(line[1], {'sum_qty': 0, 'lines': {}})
            res[line[1]]['sum_qty'] += line[0]

            if line[3] not in res[line[1]]['lines']:
                res[line[1]]['lines'][line[3]] = {
                    'qty': 0,
                    'batch_id': line[3],
                    'expiry_date': line[2] or '',
                    'location_ids': {},
                }
            res[line[1]]['lines'][line[3]]['qty'] += line[0]
            res[line[1]]['lines'][line[3]]['location_ids'].setdefault(line[4], 0.00)
            res[line[1]]['lines'][line[3]]['location_ids'][line[4]] += line[0]

        # fetch bn and product data
        for bn in self.pool.get('stock.production.lot').read(self.cr, self.uid, list(all_bn_ids.keys()), ['name'], context=context):
            bn_data[bn['id']] = bn['name']

        if full_prod_list:
            product_ids_to_fetch = list(set(list(values.get('product_ids', []))+list(all_product_ids.keys())+full_prod_list))
        else:
            product_ids_to_fetch = list(all_product_ids.keys())

        cost_price_at_date = {}
        if report.stock_level_date and product_ids_to_fetch:
            self.cr.execute("""select distinct on (product_id) product_id, old_standard_price
                from standard_price_track_changes
                where
                    product_id in %s and
                    create_date >= %s
                order by product_id, create_date asc""", (tuple(product_ids_to_fetch), values['stock_level_date']))

            for x in self.cr.fetchall():
                cost_price_at_date[x[0]] = x[1]

        for product in prod_obj.browse(self.cr, self.uid, product_ids_to_fetch, fields_to_fetch=['default_code', 'uom_id', 'name', 'standard_price', 'mml_status', 'msl_status'], context=context):
            product_data[product.id] = product
            if product.id not in res:
                res[product.id] = {
                    'sum_qty': 0,
                    'lines': {
                        '': {
                            'qty': 0,
                            'batch_id': False,
                            'expiry_date': '',
                            'location_ids': {}
                        }
                    }
                }

        # replace db id by name, code
        final_result = {}
        total_value = 0
        nb_items = 0
        status_map = {
            'F': _('No'),
            'T': _('Yes'),
            'na': '',
        }
        for product_id in res:
            product_code = product_data[product_id].default_code
            cost_price = cost_price_at_date.get(product_id, product_data[product_id].standard_price)
            rounded_qty = round(res[product_id]['sum_qty'], 6)
            final_result[product_code] = {
                'sum_qty': rounded_qty,
                'product_code': product_code,
                'product_name': product_data[product_id].name,
                'uom': product_data[product_id].uom_id.name,
                'sum_value':  cost_price * rounded_qty,
                'with_zero': with_zero,
                'moves_in_months': product_id in date_prod_list,
                'mml_status': status_map.get(product_data[product_id].mml_status, ''),
                'msl_status': status_map.get(product_data[product_id].msl_status, ''),
                'lines': {},
            }
            total_value += final_result[product_code]['sum_value']
            if rounded_qty > 0:
                nb_items += 1
            for batch_id in res[product_id]['lines']:
                rounded_batch_qty = round(res[product_id]['lines'][batch_id]['qty'], 6)
                # US-9727: With a product list, do not display empty BN/ED lines
                if not report.product_list_id or rounded_batch_qty != 0:
                    final_result[product_code]['lines'][batch_id] = {
                        'batch': bn_data.get(batch_id, ''),
                        'expiry_date': res[product_id]['lines'][batch_id]['expiry_date'],
                        'qty': rounded_batch_qty,
                        'value': cost_price * rounded_batch_qty,
                        'location_ids': dict([(x, round(y, 6)) for x, y in res[product_id]['lines'][batch_id]['location_ids'].items()]),
                    }
            # US-9727: With a product list, add 1 empty with no data if there is none found/all have been removed
            if report.product_list_id and not final_result[product_code].get('lines'):
                final_result[product_code]['lines'][False] = {
                    'batch': '',
                    'expiry_date': False,
                    'qty': 0,
                    'value': 0,
                    'location_ids': {},
                }

        fres = []
        for k in sorted(final_result.keys()):
            fres.append(final_result[k])
        return total_value, nb_items, fres

    def generate(self, context=None):
        if context is None:
            context = {}

        report = self.pool.get('export.report.stock.inventory').browse(self.cr, self.uid, self.ids[0], context=context)
        sheet = self.workbook.active

        sheet.column_dimensions['A'].width = 23.0
        sheet.column_dimensions['B'].width = 60.0
        sheet.column_dimensions['C'].width = 18.0
        sheet.column_dimensions['D'].width = 10.0
        sheet.column_dimensions['E'].width = 10.0
        sheet.column_dimensions['F'].width = 20.0
        sheet.column_dimensions['G'].width = 15.0
        ite = 8
        loc_list = []
        locations = self.get_locations(context=context)
        if not report.location_id:
            for loc in locations:
                sheet.column_dimensions[get_column_letter(ite)].width = 15.0
                loc_list.append(_('Qty: %s') % loc.name)
                ite += 1
        else:
            sheet.column_dimensions[get_column_letter(ite)].width = 15.0
            ite += 1
            sheet.column_dimensions[get_column_letter(ite)].width = 15.0
            ite += 1
            sheet.column_dimensions[get_column_letter(ite)].width = 15.0
            ite += 1
        sheet.column_dimensions[get_column_letter(ite)].width = 15.0

        # Styles
        orange_style = self.create_style_from_template('orange_style', 'A1')
        grey_style = self.create_style_from_template('grey_style', 'C1')
        grey_date_style = self.create_style_from_template('grey_date_style', 'C2')
        dark_grey_style = self.create_style_from_template('dark_grey_style', 'S13')
        default_style = self.create_style_from_template('default_style', 'A13')
        date_style = self.create_style_from_template('date_style', 'G13')

        # If the title is > 31 chars, it causes a minor error during generation and while using the file
        if loc_list:
            sheet.title = _('Export Inv. Level (all Loc.)')
        else:
            sheet.title = _('Export Inv. Level')
        # Empty cells
        empty_cell = WriteOnlyCell(sheet)
        empty_cell.style = default_style
        empty_orange_cell = WriteOnlyCell(sheet)
        empty_orange_cell.style = orange_style
        empty_grey_cell = WriteOnlyCell(sheet)
        empty_grey_cell.style = grey_style
        empty_grey_date_cell = WriteOnlyCell(sheet)
        empty_grey_date_cell.style = grey_date_style

        # Header data
        cell_h1 = WriteOnlyCell(sheet, value=_('DB/instance name'))
        cell_h1.style = orange_style
        cell_h1d = WriteOnlyCell(sheet, value=report.company_id.name)
        cell_h1d.style = grey_style
        sheet.append([cell_h1, empty_orange_cell, cell_h1d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A1:B1")
        sheet.merged_cells.ranges.append("C1:E1")

        cell_h2 = WriteOnlyCell(sheet, value=loc_list and _('Report Generation date') or _('Generated on'))
        cell_h2.style = orange_style
        cell_h2d = WriteOnlyCell(sheet, value=self.to_datetime(report.name))
        cell_h2d.style = grey_date_style
        sheet.append([cell_h2, empty_orange_cell, cell_h2d, empty_grey_date_cell, empty_grey_date_cell])
        sheet.merged_cells.ranges.append("A2:B2")
        sheet.merged_cells.ranges.append("C2:E2")

        cell_h3 = WriteOnlyCell(sheet, value=_('Stock Level date'))
        cell_h3.style = orange_style
        cell_h3d = WriteOnlyCell(sheet, value=self.to_datetime(report.stock_level_date))
        cell_h3d.style = grey_date_style
        sheet.append([cell_h3, empty_orange_cell, cell_h3d, empty_grey_date_cell, empty_grey_date_cell])
        sheet.merged_cells.ranges.append("A3:B3")
        sheet.merged_cells.ranges.append("C3:E3")

        cell_h4 = WriteOnlyCell(sheet, value=_('Specific product'))
        cell_h4.style = orange_style
        cell_h4d = WriteOnlyCell(sheet, value=report.product_id and report.product_id.default_code or '')
        cell_h4d.style = grey_style
        sheet.append([cell_h4, empty_orange_cell, cell_h4d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A4:B4")
        sheet.merged_cells.ranges.append("C4:E4")

        cell_h5 = WriteOnlyCell(sheet, value=_('Specific location'))
        cell_h5.style = orange_style
        cell_h5d = WriteOnlyCell(sheet, value=report.location_id and report.location_id.name or _('All Locations'))
        cell_h5d.style = grey_style
        sheet.append([cell_h5, empty_orange_cell, cell_h5d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A5:B5")
        sheet.merged_cells.ranges.append("C5:E5")

        cell_h6 = WriteOnlyCell(sheet, value=_('Specific batch number'))
        cell_h6.style = orange_style
        cell_h6d = WriteOnlyCell(sheet, value=report.prodlot_id and report.prodlot_id.name or '')
        cell_h6d.style = grey_style
        sheet.append([cell_h6, empty_orange_cell, cell_h6d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A6:B6")
        sheet.merged_cells.ranges.append("C6:E6")

        cell_h7 = WriteOnlyCell(sheet, value=_('Specific expiry date'))
        cell_h7.style = orange_style
        cell_h7d = WriteOnlyCell(sheet, value=self.to_datetime(report.expiry_date))
        cell_h7d.style = grey_date_style
        sheet.append([cell_h7, empty_orange_cell, cell_h7d, empty_grey_date_cell, empty_grey_date_cell])
        sheet.merged_cells.ranges.append("A7:B7")
        sheet.merged_cells.ranges.append("C7:E7")

        cell_h8 = WriteOnlyCell(sheet, value=_('Specific Product list'))
        cell_h8.style = orange_style
        cell_h8d = WriteOnlyCell(sheet, value=report.product_list_id and report.product_list_id.name or '')
        cell_h8d.style = grey_style
        sheet.append([cell_h8, empty_orange_cell, cell_h8d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A8:B8")
        sheet.merged_cells.ranges.append("C8:E8")

        cell_h9 = WriteOnlyCell(sheet, value=_('Specific Product family'))
        cell_h9.style = orange_style
        cell_h9d = WriteOnlyCell(sheet, value=report.nomen_family_id and report.nomen_family_id.complete_name or '')
        cell_h9d.style = grey_style
        sheet.append([cell_h9, empty_orange_cell, cell_h9d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A9:B9")
        sheet.merged_cells.ranges.append("C9:E9")

        cell_h10 = WriteOnlyCell(sheet, value=_('MML'))
        cell_h10.style = orange_style
        cell_h10d = WriteOnlyCell(sheet, value=report.mml_id and report.mml_id.name or '')
        cell_h10d.style = grey_style
        sheet.append([cell_h10, empty_orange_cell, cell_h10d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A10:B10")
        sheet.merged_cells.ranges.append("C10:E10")

        cell_h11 = WriteOnlyCell(sheet, value=_('MSL'))
        cell_h11.style = orange_style
        cell_h11d = WriteOnlyCell(sheet, value=report.msl_id and report.msl_id.instance_name or '')
        cell_h11d.style = grey_style
        sheet.append([cell_h11, empty_orange_cell, cell_h11d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A11:B11")
        sheet.merged_cells.ranges.append("C11:E11")

        cell_h12 = WriteOnlyCell(sheet, value=_('Including products with stock <= 0 with movements in the last months'))
        cell_h12.style = orange_style
        cell_h12d = WriteOnlyCell(sheet, value=report.display_0 and report.in_last_x_months + _(' months') or _('N/A'))
        cell_h12d.style = grey_style
        sheet.append([cell_h12, empty_orange_cell, cell_h12d, empty_grey_cell, empty_grey_cell])
        sheet.merged_cells.ranges.append("A12:B12")
        sheet.merged_cells.ranges.append("C12:E12")

        total_value, nb_items, lines = self.get_lines(report.id, context=context)
        # Lines Header data
        cell_lh_13_b = WriteOnlyCell(sheet, value=_('Number of items with stock > 0'))
        cell_lh_13_b.style = orange_style
        cell_lh_13_c = WriteOnlyCell(sheet, value=_('Total value of the generated report'))
        cell_lh_13_c.style = orange_style
        cell_lh_13_d = WriteOnlyCell(sheet, value=_('Currency'))
        cell_lh_13_d.style = orange_style
        orange_row_header = [empty_orange_cell, cell_lh_13_b, cell_lh_13_c, cell_lh_13_d, empty_orange_cell,
                             empty_orange_cell, empty_orange_cell, empty_orange_cell, empty_orange_cell,
                             empty_orange_cell, empty_orange_cell]
        if loc_list:
            or_extend = []
            oi = 1
            while oi < len(loc_list) - 2:
                or_extend.append(empty_orange_cell)
                oi += 1
            orange_row_header.extend(or_extend)
        sheet.append(orange_row_header)

        cell_lh_14_b = WriteOnlyCell(sheet, value=nb_items)
        cell_lh_14_b.style = default_style
        cell_lh_14_c = WriteOnlyCell(sheet, value=total_value)
        cell_lh_14_c.style = default_style
        cell_lh_14_d = WriteOnlyCell(sheet, value=report.company_id.currency_id.name)
        cell_lh_14_d.style = default_style
        empty_row_header = [empty_cell, cell_lh_14_b, cell_lh_14_c, cell_lh_14_d, empty_cell, empty_cell, empty_cell,
                            empty_cell, empty_cell, empty_cell, empty_cell]
        if loc_list:
            em_extend = []
            ei = 1
            while ei < len(loc_list) - 2:
                em_extend.append(empty_cell)
                ei += 1
            empty_row_header.extend(em_extend)
        sheet.append(empty_row_header)

        row_headers = [
            (_('Product Code')),
            (_('Product Description')),
            (_('UoM')),
            (_('MML')),
            (_('MSL')),
            (_('Batch')),
            (_('Exp Date')),
        ]
        if loc_list:
            row_headers.extend(loc_list)
            row_headers.append(_('Total Qty'))
        else:
            row_headers.extend([_('Qty'), _('Value'), _('Total Qty'), _('Total Value')])
        row_header = []
        for header in row_headers:
            cell_t = WriteOnlyCell(sheet, value=header)
            cell_t.style = orange_style
            row_header.append(cell_t)
        sheet.append(row_header)

        # Lines
        if loc_list:
            for prd in lines:
                for line in iter(prd['lines'].values()):
                    self.rows = []
                    if line['qty'] or (prd['with_zero'] and prd['moves_in_months']):
                        self.add_cell(prd['product_code'], default_style)
                        self.add_cell(prd['product_name'], default_style)
                        self.add_cell(prd['uom'], default_style)
                        self.add_cell(prd['mml_status'], default_style)
                        self.add_cell(prd['msl_status'], default_style)
                        self.add_cell(line['batch'], default_style)
                        self.add_cell(self.to_datetime(line['expiry_date']), date_style)
                        for loc in locations:
                            self.add_cell(line['location_ids'].get(loc.id, 0.00), default_style)
                        self.add_cell(line.get('qty', 0.00), default_style)
                        sheet.append(self.rows)
        else:
            for prd in lines:
                self.rows = []
                if prd['sum_qty'] or prd['with_zero']:
                    self.add_cell(prd['product_code'], dark_grey_style)
                    self.add_cell(prd['product_name'], dark_grey_style)
                    self.add_cell(prd['uom'], dark_grey_style)
                    self.add_cell(prd['mml_status'], dark_grey_style)
                    self.add_cell(prd['msl_status'], dark_grey_style)
                    self.add_cell('', dark_grey_style)
                    self.add_cell('', dark_grey_style)
                    self.add_cell('', dark_grey_style)
                    self.add_cell('', dark_grey_style)
                    self.add_cell(round(prd['sum_qty'], 2), dark_grey_style)
                    self.add_cell(round(prd['sum_value'], 2), dark_grey_style)
                    sheet.append(self.rows)
                    for line in iter(prd['lines'].values()):
                        self.rows = []
                        if line['qty'] or prd['with_zero']:
                            self.add_cell(prd['product_code'], default_style)
                            self.add_cell(prd['product_name'], default_style)
                            self.add_cell(prd['uom'], default_style)
                            self.add_cell(prd['mml_status'], default_style)
                            self.add_cell(prd['msl_status'], default_style)
                            self.add_cell(line['batch'], default_style)
                            self.add_cell(self.to_datetime(line['expiry_date']), date_style)
                            self.add_cell(round(line['qty'], 2), default_style)
                            self.add_cell(round(line['value'], 2), default_style)
                            self.add_cell('', default_style)
                            self.add_cell('', default_style)
                            sheet.append(self.rows)


XlsxReport('report.report_stock_inventory_xlsx', parser=export_report_stock_inventory_parser, template='addons/specific_rules/report/report_stock_inventory.xlsx')
