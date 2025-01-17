# -*- coding: utf-8 -*-

from tools.misc import Path, file_open
from report.interface import report_int
import netsvc
import openpyxl
from tempfile import NamedTemporaryFile
from copy import copy
import pooler
import datetime
from openpyxl.styles import NamedStyle
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles.protection import Protection
from openpyxl.styles import Alignment


class XlsxReport(report_int):
    def __init__(self, name, parser, write_only=True, template=False):
        """
            Manage xlsx document
            Recommended usage:
                write_only=True, to create a memory improved workbook, but with some restrictions:
                    https://openpyxl.readthedocs.io/en/stable/optimized.html#write-only-mode

            :param template: path to an existing xlsx file
                if write_only == True: the template can be used to copy styles, row height, column width
                if write_only == False: a copy of the template is returned, ready for reading and writing


        """

        super(XlsxReport, self).__init__(name)
        self.parser = parser
        self.write_only = write_only
        self.template = template
        if not netsvc.Service._services.get(self.name2):
            netsvc.Service._services[self.name2] = parser

    def create(self, cr, uid, ids, datas, context):
        wb_t = False
        if self.template:
            if self.write_only:
                wb_t = openpyxl.load_workbook(file_open(self.template, mode='rb'))
            else:
                wb = openpyxl.load_workbook(file_open(self.template, mode='rb'))

        if self.write_only:
            wb = openpyxl.Workbook(write_only=True)
            wb.create_sheet()
        else:
            wb = openpyxl.Workbook()
        wb.iso_dates = True
        parser = self.parser(cr, uid, ids, wb, wb_t, context)
        parser.model = datas.get('model')
        parser.generate(context=context)
        tmp = NamedTemporaryFile(delete=False)
        wb.save(tmp.name)
        wb.close()
        if wb_t:
            wb_t.close()
        tmp.seek(0)
        return (Path(tmp.name, delete=True), 'xlsx')

class XlsxReportParser():
    def __init__(self, cr, uid, ids, workbook, workbook_template, context):
        self.cr = cr
        self.uid = uid
        self.ids = ids
        self.context = context
        self.workbook = workbook
        self.workbook_template = workbook_template
        self.pool = pooler.get_pool(cr.dbname)

    def duplicate_column_dimensions(self, default_width=None):
        assert self.workbook_template, "duplicate_column_dimensions can be used only with a tempate"

        for x in range(1, self.workbook_template.active.max_column+1):
            letter = get_column_letter(x)
            tmp_column = self.workbook_template.active.column_dimensions[letter]
            width = False
            if not tmp_column.customWidth and default_width:
                width = default_width
            elif tmp_column.customWidth:
                width = tmp_column.width

            if width:
                self.workbook.active.column_dimensions[letter].width = width

    def duplicate_row_dimensions(self, row_range):
        assert self.workbook_template, "duplicate_row_dimensions can be used only with a tempate"

        for x in row_range:
            self.workbook.active.row_dimensions[x].height = self.workbook_template.active.row_dimensions[x].height

    def apply_template_style(self, cell_index, target):
        """
            :param cell_index: index of the cell, example A1
            :param target: cell object
        """
        self.duplicate_cell_style(self.workbook_template.active[cell_index], target)

    def duplicate_cell_style(self, src, target):
        """
            :param src: cell object
            :param target: object cell or workbook
        """
        target.style = copy(src.style)
        target.font = copy(src.font)
        target.fill = copy(src.fill)
        target.border = copy(src.border)
        target.alignment = copy(src.alignment)
        target.number_format = copy(src.number_format)

    def create_style_from_cell(self, name, src):
        """
            :param src: cell object
        """
        new_style = NamedStyle(name=name)
        self.duplicate_cell_style(src, new_style)
        return new_style

    def create_style_from_template(self, name, cell_index):
        """
            :param cell_index: index of the cell, example A1
        """
        new_style = self.create_style_from_cell(name, self.workbook_template.active[cell_index])
        self.workbook.add_named_style(new_style)
        return new_style

    def getSel(self, o, field):
        return self.pool.get('ir.model.fields').get_browse_selection(self.cr, self.uid, o, field, self.context)

    def to_datetime(self, value):
        if not value:
            return ''
        if len(value) == 10:
            str_date = value + ' 00:00:00'
        else:
            str_date = value
        return datetime.datetime.strptime(str_date, "%Y-%m-%d %H:%M:%S")

    def cell_ro(self, value, style=None, unlock=None, copy_style=None, wrap_text=None):
        new_cell = WriteOnlyCell(self.workbook.active, value=value)
        if style:
            new_cell.style = style
        elif copy_style:
            copied_style_name = '%s_style' % copy_style
            if copied_style_name in self.workbook.named_styles:
                new_cell.style = copied_style_name
            else:
                new_cell.style = self.create_style_from_template(copied_style_name, copy_style)
        if unlock:
            new_cell.protection = Protection(locked=False)
        if wrap_text:
            new_cell.alignment = new_cell.alignment.copy(wrapText=True)
        return new_cell


