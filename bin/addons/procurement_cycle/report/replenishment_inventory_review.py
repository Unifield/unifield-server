# -*- coding: utf-8 -*-
from spreadsheet_xml.xlsx_write import XlsxReport
from spreadsheet_xml.xlsx_write import XlsxReportParser
from datetime import datetime
from dateutil.relativedelta import relativedelta
from tools import misc
from tools.translate import _
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.styles import NamedStyle, Alignment, Font, Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.cell import WriteOnlyCell


class inventory_parser(XlsxReportParser):

    def add_cell(self, value=None, style=None):
        new_cell = WriteOnlyCell(self.workbook.active, value=value)
        if style:
            new_cell.style = style
        self.rows.append(new_cell)

    def generate(self, context=None):
        inventory = self.pool.get('replenishment.inventory.review').browse(self.cr, self.uid, self.ids[0], context=context)

        workbook = self.workbook
        sheet = workbook.active
        sheet.freeze_panes = 'D16'

        default_alignment = Alignment(wrapText=True, horizontal='center',  vertical='center')
        default_border =  Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
        )
        default_format = {
            'alignment': default_alignment,
            'font': Font(name='Calibri', sz=11.0),
            'border': default_border,
        }
        title_style = NamedStyle(name='title', font=Font(name='Calibri', sz=24.0, bold=True), fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDEE2')), alignment=default_alignment)
        sub_title_style = NamedStyle(
            name='sub_title',
            font=Font(name='Calibri', sz=11.0, bold=True),
            fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFDCDEE2')),
            alignment=Alignment(wrapText=True, horizontal='left', vertical='center'),
            border=default_border
        )
        sub_value_style = NamedStyle(
            name='sub_value',
            font=Font(name='Calibri', sz=11.0),
            alignment=default_alignment,
            border=default_border
        )
        main_header_style = NamedStyle(name='main_header_style', font=Font(name='Calibri', sz=11.0, bold=True), alignment=default_alignment, border=default_border)
        green_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF00B050'))
        red_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFF0000'))
        orange_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFFC000'))
        blue_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF00B0F0'))
        yellow_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFFFF00'))
        pink1_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFF88CF0'))
        pink2_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFD2AED6'))

        date_style = NamedStyle(name='date', number_format='DD/MM/YYYY', **default_format)
        default_style = NamedStyle(name='default', number_format='0.00', **default_format)
        precent_style = NamedStyle(name='precent', number_format='0.00%',  **default_format)
        float_style = NamedStyle(name='float', number_format='0.00',  **default_format)
        grey_style = NamedStyle(name='grey', fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFD9D9D9')),  **default_format)

        sheet.row_dimensions[1].height = 33.75
        sheet.row_dimensions[2].height = 31.5
        sheet.row_dimensions[3].height = 15.0
        sheet.row_dimensions[4].height = 15.0
        sheet.row_dimensions[5].height = 15.0
        sheet.row_dimensions[6].height = 15.0
        sheet.row_dimensions[7].height = 15.0
        sheet.row_dimensions[8].height = 15.0
        sheet.row_dimensions[9].height = 15.0
        sheet.row_dimensions[10].height = 15.0
        sheet.row_dimensions[11].height = 15.0
        sheet.row_dimensions[12].height = 15.0
        sheet.row_dimensions[13].height = 15.0
        sheet.row_dimensions[14].height = 15.0
        sheet.row_dimensions[15].height = 75.0
        sheet.column_dimensions['A'].width = 15.28515625
        sheet.column_dimensions['B'].width = 38.85546875
        sheet.column_dimensions['C'].width = 20.5703125
        sheet.column_dimensions['D'].width = 23.7109375
        sheet.column_dimensions['E'].width = 32.5703125
        sheet.column_dimensions['F'].width = 17.85546875
        sheet.column_dimensions['G'].width = 21.7109375
        sheet.column_dimensions['H'].width = 10.75
        sheet.column_dimensions['I'].width = 10.75
        sheet.column_dimensions['J'].width = 11.28515625
        sheet.column_dimensions['K'].width = 10.75
        sheet.column_dimensions['L'].width = 10.75
        sheet.column_dimensions['M'].width = 11.28515625
        sheet.column_dimensions['N'].width = 10.75
        sheet.column_dimensions['O'].width = 10.75
        sheet.column_dimensions['P'].width = 10.75
        sheet.column_dimensions['Q'].width = 10.75
        sheet.column_dimensions['R'].width = 10.75
        sheet.column_dimensions['S'].width = 10.75
        sheet.column_dimensions['T'].width = 11.5703125
        sheet.column_dimensions['U'].width = 11.28515625
        sheet.column_dimensions['V'].width = 11.28515625
        sheet.column_dimensions['W'].width = 10.75
        sheet.column_dimensions['X'].width = 15.0
        sheet.column_dimensions['Y'].width = 14.5703125
        sheet.column_dimensions['Z'].width = 10.75
        sheet.column_dimensions['AA'].width = 10.75
        sheet.column_dimensions['AB'].width = 10.75
        sheet.column_dimensions['AC'].width = 10.75
        sheet.column_dimensions['AD'].width = 10.75
        sheet.column_dimensions['AE'].width = 10.75
        sheet.column_dimensions['AF'].width = 10.75
        sheet.column_dimensions['AG'].width = 10.75
        sheet.column_dimensions['AH'].width = 10.75
        sheet.column_dimensions['AI'].width = 19.0
        sheet.column_dimensions['AJ'].width = 10.75
        sheet.column_dimensions['AK'].width = 10.75
        sheet.column_dimensions['AL'].width = 10.75
        sheet.column_dimensions['AM'].width = 10.75
        sheet.column_dimensions['AN'].width = 10.75
        sheet.column_dimensions['AO'].width = 10.75
        sheet.column_dimensions['AP'].width = 10.75
        sheet.column_dimensions['AQ'].width = 10.75
        sheet.column_dimensions['AR'].width = 11.140625
        sheet.column_dimensions['AS'].width = 10.75


        sheet.column_dimensions.group('AR', get_column_letter(column_index_from_string('AR')+2*inventory.projected_view - 1), hidden=False)
        sheet.column_dimensions.group('D', 'E', hidden=False)
        sheet.column_dimensions.group('V', 'Y', hidden=False)
        sheet.row_dimensions.group(2, 14, hidden=False)


        sheet.title = _('Inv. Review')
        cell_title = WriteOnlyCell(sheet, value=_('Inventory Review'))
        cell_title.style = title_style
        sheet.append([cell_title, ''])
        sheet.merged_cells.ranges.append("A1:B1")


        sheet.append([])
        row_index = 3


        time_unit_str = self.getSel(inventory, 'time_unit')
        doc_headers = [
            (_('Stock location configuration description'), inventory.location_config_id.description, ''),
            (_('Scheduled reviews periodicity'), inventory.frequence_name or '', ''),
            (_('Generated'), datetime.now(), 'dt'), # TODO
            (_('RR-AMC periodicity (1st month)'), self.to_datetime(inventory.amc_first_date), 'd'),
            (_('RR-AMC periodicity (last month)'), self.to_datetime(inventory.amc_last_date), 'd'),
            (_('Projected view (months from generation date)'), inventory.projected_view, ''),
            (_('Final day of projection'), self.to_datetime(inventory.final_date_projection), 'd'),
            (_('Sleeping stock alert parameter (months)'), inventory.sleeping, ''),
            (_('Display durations in'), time_unit_str, ''),
        ]

        for title, value, v_type in doc_headers:
            cell_t = WriteOnlyCell(sheet, value=title)
            cell_t.style = sub_title_style
            cell_v = WriteOnlyCell(sheet, value=value)
            cell_v.style = sub_value_style
            if v_type == 'd':
                cell_v.number_format = 'DD/MM/YYYY'
            elif v_type == 'dt':
                cell_v.number_format = 'DD/MM/YYYY HH:MM'
            sheet.append([cell_t, '', cell_v])
            sheet.merged_cells.ranges.append("A%(row_index)d:B%(row_index)d" % {'row_index':row_index})
            row_index += 1

        sheet.append([])
        sheet.append([])
        sheet.append([])

        row_headers = [
            (_('Product Code'), green_fill),
            (_('Product Description'), green_fill),
            (_('RR Lifecycle'), green_fill),
            (_('Replaced/Replacing'), green_fill),
            (_('Primary Product list'), green_fill),
            (_('Warnings Recap'), red_fill),
            (_('Segment Ref/name'), orange_fill),
            (_('RR (threshold) applied'), orange_fill),
            (_('RR (qty) applied'), orange_fill),
            (_('Min'), orange_fill),
            (_('Max'), orange_fill),
            (_('Automatic Supply order  qty'), orange_fill),
            ('%s %s' % (_('Internal LT'), time_unit_str), orange_fill),
            ('%s %s' % (_('External LT'), time_unit_str), orange_fill),
            (_('Total lead time'), orange_fill),
            ('%s %s' % (_('Order coverage'), time_unit_str), orange_fill),
            (_('SS (Qty)'), orange_fill),
            (_('Buffer (Qty)'), orange_fill),
            (_('Valid'), orange_fill),
            (_('RR-FMC (average for period)'), orange_fill),
            (_('RR-AMC (average for AMC period)'), orange_fill),
            (_('Standard Deviation HMC'), orange_fill),
            (_('Coefficient of Variation of HMC'), orange_fill),
            (_('Standard Deviation of HMC vs FMC'), orange_fill),
            (_('Coefficient of Variant of HMC and FMC'), orange_fill),
            (_('Real Stock in location(s)'), blue_fill),
            (_('Pipeline Qty'), blue_fill),
            (_('Reserved Stock Qty'), blue_fill),
            (_('(RR-FMC)Projected stock Qty'), blue_fill),
            (_('(RR-AMC) Projected stock'), blue_fill),
            (_('Qty of Projected Expiries before consumption'), blue_fill),
            (_('Qty expiring within period'), blue_fill),
            (_('Open Loan on product (Yes/No)'), blue_fill),
            (_('Donations pending (Yes/No)'), blue_fill),
            (_('Sleeping stock Qty'), blue_fill),
            ('%s %s' % (time_unit_str,_('of supply (RR-AMC)')), blue_fill),
            ('%s %s' % (time_unit_str, _('of supply (RR-FMC)')), blue_fill),
            (_('Qty lacking before next RDD'), blue_fill),
            (_('Qty lacking needed by'), yellow_fill),
            (_('ETA date of next pipeline'), yellow_fill),
            (_('Date to start preparing the next order'),  yellow_fill),
            (_('Next order to be generated/issued by date'), yellow_fill),
            (_('RDD for next order'), yellow_fill),
        ]


        for nb_month in range(0, inventory.projected_view):
            row_headers.append(('%s M%s\n %s' % (_('RR-FMC'), nb_month, self.get_month(inventory.generation_date, nb_month)), pink1_fill))

        for nb_month in range(0, inventory.projected_view):
            row_headers.append(('%s\nM%s %s' % (_('Projected'), nb_month, self.get_month(inventory.generation_date, nb_month)), pink2_fill))

        row_header = []
        for header, bg_color in row_headers:
            cell_t = WriteOnlyCell(sheet, value=header)
            cell_t.style = main_header_style
            cell_t.fill = bg_color
            row_header.append(cell_t)
        sheet.append(row_header)

        new_row = 16
        for line in inventory.line_ids:
            self.rows = []
            sheet.row_dimensions[new_row].height = 40

            self.add_cell(line.product_id.default_code, default_style)
            self.add_cell(line.product_id.name, default_style)
            self.add_cell(line.segment_ref_name and self.getSel(line, 'status') or _('N/A'), default_style)
            self.add_cell(line.paired_product_id and line.paired_product_id.default_code or '',  default_style)
            self.add_cell(line.primay_product_list or '',  default_style)
            self.add_cell(line.warning or '', default_style)
            self.add_cell(line.segment_ref_name or '', default_style)
            self.add_cell(line.rule == 'cycle' and 'PAS' or '', default_style)
            self.add_cell(line.segment_ref_name and self.getSel(line, 'rule') or '', default_style)
            if line.rule == 'minmax':
                self.add_cell(line.min_qty, default_style)
                self.add_cell(line.max_qty, default_style)
            else:
                self.add_cell('', grey_style)
                self.add_cell('', grey_style)

            if line.rule == 'auto':
                self.add_cell(line.auto_qty, default_style)
            else:
                self.add_cell('', grey_style)

            self.add_cell(line.internal_lt, default_style) # float vs int
            self.add_cell(line.external_lt, default_style) # float vs int
            self.add_cell(line.total_lt, default_style) # float vs int
            self.add_cell(line.order_coverage, default_style) # float vs int
            self.add_cell(line.safety_stock_qty, default_style)
            if line.rule == 'cycle':
                self.add_cell(line.buffer_qty, default_style)
            else:
                self.add_cell('', grey_style)

            self.add_cell(line.valid_rr_fmc and _('Yes') or _('No'), default_style)
            self.add_cell(line.rr_fmc_avg, float_style)
            self.add_cell(line.rr_amc, float_style)
            self.add_cell(line.std_dev_hmc, float_style)
            self.add_cell(line.coef_var_hmc/100., precent_style)

            if line.rule == 'cycle':
                self.add_cell(line.std_dev_hmc_fmc, float_style)
                self.add_cell(line.coef_var_hmc_fmc, float_style)
            else:
                self.add_cell('', grey_style)
                self.add_cell('', grey_style)

            self.add_cell(line.real_stock, default_style)
            self.add_cell(line.pipeline_qty, default_style)
            self.add_cell(line.reserved_stock_qty, default_style)

            if line.projected_stock_qty and line.rule == 'cycle':
                self.add_cell(line.projected_stock_qty, default_style)
            else:
                self.add_cell('', default_style)

            if line.projected_stock_qty_amc and (line.rule == 'cycle' or not line.segment_ref_name):
                self.add_cell(line.projected_stock_qty_amc, default_style)
            else:
                self.add_cell('', default_style)

            if line.expired_qty_before_cons:
                self.add_cell(line.expired_qty_before_cons, date_style)
            else:
                self.add_cell('', default_style)

            if line.total_expired_qty:
                self.add_cell(line.total_expired_qty, date_style)
            else:
                self.add_cell('', default_style)

            self.add_cell(line.open_loan and _('Yes') or _('No'),  default_style)
            self.add_cell(line.open_donation and _('Yes') or _('No'), default_style)

            self.add_cell(line.sleeping_qty, default_style)
            self.add_cell(line.unit_of_supply_amc, default_style)

            self.add_cell(line.unit_of_supply_fmc, default_style)
            self.add_cell(line.qty_lacking, default_style)
            self.add_cell(self.to_datetime(line.qty_lacking_needed_by), date_style)
            self.add_cell(self.to_datetime(line.eta_for_next_pipeline), date_style)
            self.add_cell(self.to_datetime(line.date_preparing), date_style)
            self.add_cell(self.to_datetime(line.date_next_order_validated), date_style)
            self.add_cell(self.to_datetime(line.date_next_order_rdd), date_style)

            for detail_pas in line.pas_ids:
                if detail_pas.rr_fmc is not None and detail_pas.rr_fmc is not False:
                    self.add_cell(detail_pas.rr_fmc, default_style)
                else:
                    self.add_cell(_('Invalid FMC'), default_style)

            for detail_pas in line.pas_ids:
                if detail_pas.projected is not None and detail_pas.projected is not False:
                    self.add_cell(detail_pas.projected, default_style)
            if not line.pas_ids:
                for nb_month in range(0, inventory.projected_view):
                    self.add_cell('', default_style)
                    self.add_cell('', default_style)

            new_row += 1
            sheet.append(self.rows)

    def get_month(self, start, nb_month):
        return _(misc.month_abbr[(datetime.strptime(start, '%Y-%m-%d %H:%M:%S') + relativedelta(hour=0, minute=0, second=0, months=nb_month)).month])

XlsxReport('report.report_replenishment_inventory_review_xls', parser=inventory_parser, template='addons/procurement_cycle/report/replenishment_inventory_review.xls', write_only=True)

