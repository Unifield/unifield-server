# -*- coding: utf-8 -*-
from osv import fields
from osv import osv
from tools.translate import _
from openpyxl import load_workbook
import base64
import tools
import time
from io import BytesIO


class wizard_return_from_unit_import(osv.osv_memory):
    _name = 'wizard.return.from.unit.import'
    _rec_name = 'picking_id'

    _columns = {
        'picking_id': fields.many2one('stock.picking', string='Incoming shipment', required=True, readonly=True),
        'state': fields.selection([('draft', 'Draft'), ('in_progress', 'Import in progress'), ('error', 'Error'), ('done', 'Done')], string='State', readonly=True),
        'file_to_import': fields.binary(string='File to import', filters='*.xls*'),
        'filename': fields.char(size=64, string='Filename'),
        'message': fields.text(string='Message', readonly=True, translate=True),
    }

    def export_scratch_template_file(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids,(int,long)):
            ids = [ids]

        return {'type': 'ir.actions.report.xml', 'report_name': 'report_return_from_unit_xls', 'context': context}

    def go_to_in(self, cr, uid, ids, context=None):
        '''
        Return to the initial view.
        '''
        if isinstance(ids, (int, long)):
            ids = [ids]
        for wiz in self.browse(cr, uid, ids, context=context):
            pick = wiz.picking_id
            xmild = self.pool.get('stock.picking')._hook_picking_get_view(cr, uid, ids, context=context, pick=pick)
            res = self.pool.get('ir.actions.act_window').open_view_from_xmlid(cr, uid, xmild, ['form', 'tree'], context=context)
            res['res_id'] = pick.id
            return res

    def import_in(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids,(int,long)):
            ids = [ids]

        loc_obj = self.pool.get('stock.location')
        prod_obj = self.pool.get('product.product')
        uom_obj = self.pool.get('product.uom')
        lot_obj = self.pool.get('stock.production.lot')

        start_time = time.time()
        wiz = self.browse(cr, uid, ids[0], context=context)
        pick = wiz.picking_id
        if not wiz.file_to_import:
            raise osv.except_osv(_('Error'), _('Nothing to import.'))
        self.write(cr, uid, wiz.id, {'state': 'in_progress'}, context=context)

        wb = load_workbook(filename=BytesIO(base64.decodestring(wiz.file_to_import)), read_only=True)
        sheet = wb.active

        header_message = ''
        # From (Ext CU)
        from_loc_id = False
        if sheet['C9'].value:
            ext_cu = sheet['C9'].value
            loc_ids = loc_obj.search(cr, uid, [('name', '=ilike', ext_cu), ('usage', '=', 'customer'), ('location_category', '=', 'consumption_unit')], context=context)
            if loc_ids:
                if loc_ids[0] == pick.ext_cu.id:
                    from_loc_id = loc_ids[0]
                else:
                    header_message += _('\nThe imported Ext. C.U. must be the same as in the IN (%s).') % (pick.ext_cu.name,)
            else:
                header_message += _('\nThere is no Ext. C.U. with the name %s.') % (ext_cu,)
        else:
            header_message += _('\nYou must fill "From" with an Ext. C.U. location.')

        # To (Internal Location)
        to_loc_id = False
        if sheet['F9'].value:
            to_loc = sheet['F9'].value
            loc_ids = loc_obj.search(cr, uid, [('name', '=ilike', to_loc), ('usage', '=', 'internal'), ('location_category', 'in', ['stock', 'consumption_unit', 'eprep'])], context=context)
            if loc_ids:
                to_loc_id = loc_ids[0]
            else:
                header_message += _('\nThere is no Internal Location with the name %s.') % (to_loc,)
        else:
            header_message += _('\nYou must fill "To" with an Internal location.')

        lines = []
        lines_err = []
        lines_warn = []
        message = ''
        db_datetime_format = self.pool.get('date.tools').get_db_datetime_format(cr, uid, context=context)
        today = time.strftime(db_datetime_format)
        def_line = {'picking_id': pick.id, 'location_id': from_loc_id, 'location_dest_id': to_loc_id,
                    'reason_type_id': pick.reason_type_id.id, 'date': today, 'date_expected': today}
        for cell in sheet.iter_rows(min_row=13, min_col=1, max_col=8):
            if not cell[1].value:  # Stop looking at lines if there is no product
                break
            line = def_line.copy()
            line_err = ''
            line_warn = ''

            line_num = cell[0].row or ''

            # Product Code and BN/ED
            prod_name = cell[1].value
            prod_ids = prod_obj.search(cr, uid, [('default_code', '=ilike', prod_name)], context=context)
            if not prod_ids:
                lines.append([{}, _('Line %s: There is no active product %s. ') % (line_num, prod_name), ''])
                continue

            ftf = ['name', 'list_price', 'uom_id', 'perishable', 'batch_management']
            prod = prod_obj.browse(cr, uid, prod_ids[0], fields_to_fetch=ftf, context=context)
            line.update({'product_id': prod.id, 'name': prod.name, 'price_unit': prod.list_price})

            # BN/ED
            if prod.batch_management or prod.perishable:
                bn_name = cell[5].value
                ed = cell[6].value
                if prod.perishable:
                    if ed and (cell[6].data_type != 'd' or not cell[6].is_date):
                        line_warn += _('%s: The Expiry Date must be a date.') % (prod_name, )
                        ed = False
                    if not prod.batch_management and bn_name:
                        line_warn += _("%s: a batch number is defined on the imported file but the product doesn't require batch number - Batch ignored") % (prod_name, )
                        bn_name = False
                    if ed:
                        try:
                            ed = ed.strftime('%Y-%m-%d')  # Fix format
                        except ValueError:
                            line_warn += _('%s the Expiry Date %s is not correct.') % (prod_name, ed)

                if (bn_name and ed) or ed:
                    bn_id = lot_obj._get_or_create_lot(cr, uid, bn_name, ed, prod.id, context=context)
                    line.update({'prodlot_id': bn_id, 'expired_date': ed})

            # Quantity
            qty = cell[3].value
            if qty:
                if cell[3].data_type == 'n':
                    line.update({'product_qty': qty})
                else:
                    try:
                        qty = float(qty.rstrip().replace(',', '.'))
                        line.update({'product_qty': qty})
                    except ValueError:
                        line_err += _('The Quantity must be a number. ')
            else:
                line_err += _('The Quantity is mandatory for each line. ')

            # UoM
            uom_name = cell[4].value
            if uom_name:
                uom_ids = uom_obj.search(cr, uid, [('name', '=ilike', uom_name)], context=context)
                if uom_ids:
                    uom_id = uom_ids[0]
                    # Check the uom category consistency
                    if prod and not self.pool.get('uom.tools').check_uom(cr, uid, prod.id, uom_id, context):
                        uom_id = prod.uom_id.id
                        line_err += _('The UoM imported was not in the same category than the UoM of the product. The UoM of the product was taken instead. ')
                    line.update({'product_uom': uom_id})
            else:
                line_err += _('The UoM is mandatory for each line. ')

            # Comment
            if cell[7].value:
                line.update({'comment': tools.ustr(cell[7].value)})

            if line_err:
                line_err = _('Line %s: ') % (line_num, ) + line_err
            if line_warn:
                line_warn = _('Line %s: ') % (line_num, ) + line_warn
            lines.append([line, line_err, line_warn])

        wiz_state = 'done'
        imp_lines = 0
        for line, line_err, line_warn in lines:
            if not line_err and not header_message:
                self.pool.get('stock.move').create(cr, uid, line, context=context)
                imp_lines += 1
            elif line_err:
                lines_err.append(line_err)
            if line_warn:
                lines_warn.append(line_warn)
        if lines_err or header_message:
            if lines_err:
                message = '%s:\n%s' % (_('Errors'), "\n".join(lines_err))
            wiz_state = 'error'

        end_time = time.time()
        total_time = str(round(end_time - start_time)) + _(' second(s)')
        final_message = _(''' 
Importation completed in %s!
# of imported lines : %s on %s lines
# of ignored lines: %s
# of errors to correct: %s

%s
%s''') % (total_time, imp_lines, len(lines), len(lines) - imp_lines, len(lines_err), header_message, message)
        if lines_warn:
            final_message += "\n%s:\n%s" % (_('Warning'), "\n".join(lines_warn))

        self.write(cr, uid, wiz.id, {'state': wiz_state, 'message': final_message}, context=context)

        wb.close()  # Close manually because of readonly
        return True


wizard_return_from_unit_import()
