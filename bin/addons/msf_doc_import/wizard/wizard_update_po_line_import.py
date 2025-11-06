# -*- coding: utf-8 -*-
from osv import fields
from osv import osv
from tools.translate import _
from openpyxl import load_workbook
import base64
import tools
import time
from io import BytesIO
from datetime import datetime


class wizard_update_po_line_import(osv.osv_memory):
    _name = 'wizard.update.po.line.import'
    _rec_name = 'po_id'

    _columns = {
        'po_id': fields.many2one('purchase.order', string='Purchase Order', required=True, readonly=True),
        'state': fields.selection([('draft', 'Draft'), ('in_progress', 'Import in progress'), ('error', 'Error'), ('done', 'Done')], string='State', readonly=True),
        'file_to_import': fields.binary(string='File to import', filters='*.xls*'),
        'filename': fields.char(size=64, string='Filename'),
        'message': fields.text(string='Message', readonly=True, translate=True),
    }

    def export_po_update_template_file(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        filename = _('Update PO lines Export')
        if ids:
            po = self.browse(cr, uid, ids[0], fields_to_fetch=['po_id'], context=context).po_id
            filename += ' ' + po.name
        filename += ' ' + datetime.today().strftime('%d-%m-%Y_%H%M')

        return {
            'type': 'ir.actions.report.xml',
            'report_name': 'report_update_po_line_export',
            'datas': {'target_filename': filename},
            'context': context
        }

    def go_to_po(self, cr, uid, ids, context=None):
        '''
        Return to the initial view.
        '''
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]
        for wiz_obj in self.read(cr, uid, ids, ['po_id'], context=context):
            po_id = wiz_obj['po_id']
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'purchase.order',
                'view_type': 'form',
                'view_mode': 'form, tree',
                'target': 'crush',
                'res_id': po_id,
                'context': context,
            }
        return True

    def best_matching_line(self, cr, uid, po_id, used_pol_ids, line_num, prod_id, comment, qty, context=None):
        """
        Try to find the best matching line in a PO for an imported line
        """
        if context is None:
            context = {}

        sql = """SELECT id FROM purchase_order_line 
        WHERE order_id = %(po_id)s AND line_number = %(line_num)s AND state NOT IN ('cancel', 'cancel_r')"""
        sql_cond = {'po_id': po_id, 'line_num': line_num}

        sql_wheres = []
        if prod_id and comment and qty:
            sql_wheres.append("(product_id = %(prod_id)s AND comment = %(comment)s AND product_qty = %(qty)s)")
        if prod_id and comment:
            sql_wheres.append("(product_id = %(prod_id)s AND comment = %(comment)s)")
        if prod_id and qty:
            sql_wheres.append("(product_id = %(prod_id)s AND product_qty = %(qty)s)")
        if comment and qty:
            sql_wheres.append("(comment = %(comment)s AND product_qty = %(qty)s)")
        if prod_id:
            sql_wheres.append("product_id = %(prod_id)s")
            sql_cond['prod_id'] = prod_id
        if comment:
            sql_wheres.append("comment = %(comment)s")
            sql_cond['comment'] = comment
        if qty:
            sql_wheres.append("product_qty = %(qty)s")
            sql_cond['qty'] = qty

        sql_used_ids, sql_mod = '', ''
        if used_pol_ids:
            if len(used_pol_ids) == 1:
                sql_used_ids = ' AND id != %(used_pol_ids)s'
                sql_cond['used_pol_ids'] = used_pol_ids[0]
            else:
                sql_used_ids = ' AND id NOT IN %(used_pol_ids)s'
                sql_cond['used_pol_ids'] = tuple(used_pol_ids)
        if sql_wheres:
            sql_mod = ' AND (' + ' OR '.join(sql_wheres) + ')'

        sql = sql + sql_used_ids + sql_mod + """ ORDER BY id LIMIT 1""" # not_a_user_entry
        cr.execute(sql, sql_cond)

        pol_ids = cr.fetchone()
        if pol_ids:
            return pol_ids[0]

        # If no match found, use the domain without modifiers
        pol_obj = self.pool.get('purchase.order.line')
        def_domain = [('order_id', '=', po_id), ('id', 'not in', used_pol_ids), ('line_number', '=', line_num)]
        matching_ids = pol_obj.search(cr, uid, def_domain, context=context)
        if matching_ids:
            return matching_ids[0]

        return False

    def import_file(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        prod_obj = self.pool.get('product.product')
        pol_obj = self.pool.get('purchase.order.line')
        fields_obj = self.pool.get('ir.model.fields')

        start_time = time.time()
        wiz = self.browse(cr, uid, ids[0], context=context)
        po = wiz.po_id
        if not wiz.file_to_import:
            raise osv.except_osv(_('Error'), _('Nothing to import.'))
        self.write(cr, uid, wiz.id, {'state': 'in_progress'}, context=context)

        wb = load_workbook(filename=BytesIO(base64.b64decode(wiz.file_to_import)), read_only=True)
        sheet = wb.active

        lines = []
        lines_err = []
        message = ''
        used_pol_ids = []
        for cell in sheet.iter_rows(min_row=2, min_col=0, max_col=7):
            line_num = cell[0].value
            if not line_num:  # Stop looking at lines if there is no line_number
                break
            pol_data = {}
            line_err = ''

            row_num = cell[0].row or ''

            # Product
            prod_name = cell[1].value
            prod_id = False
            if prod_name:
                prod_ids = prod_obj.search(cr, uid, [('default_code', '=ilike', prod_name)], context=context)
                if not prod_ids:
                    line_err += _('There is no active product %s. ') % (prod_name, )
                else:
                    prod_id = prod_ids[0]

            # Comment
            comment = cell[3].value
            if comment:
                comment = tools.ustr(comment)
            elif not prod_name:
                line_err += _('The Comment is mandatory if there is no Product Code. ')
            pol_data.update({'comment': comment or ''})

            # Quantity
            qty = cell[4].value
            if qty:
                if cell[4].data_type != 'n':
                    try:
                        qty = float(qty.rstrip().replace(',', '.'))
                    except ValueError:
                        line_err += _('The Quantity must be a number. ')
                if isinstance(qty, float) or isinstance(qty, int):
                    if qty < 0:
                        line_err += _('The Quantity must be a positive. ')
                    else:
                        pol_data.update({'product_qty': qty})
            else:
                line_err += _('The Quantity is mandatory for each line. ')

            # Unit Price
            price = cell[6].value
            if price:
                if cell[6].data_type != 'n':
                    try:
                        price = float(price.rstrip().replace(',', '.'))
                    except ValueError:
                        line_err += _('The Unit Price must be a number. ')
                if isinstance(price, float) or isinstance(price, int):
                    if price < 0:
                        line_err += _('The Unit Price must be positive. ')
                    else:
                        pol_data.update({'price_unit': price})
            else:
                line_err += _('The Unit Price is mandatory for each line. ')

            if line_err:
                pol_id = False
            else:
                pol_id = self.best_matching_line(cr, uid, po.id, used_pol_ids, line_num, prod_id, comment, qty, context=context)
                if pol_id:
                    used_pol_ids.append(pol_id)
                    pol_state = pol_obj.read(cr, uid, pol_id, ['state'], context=context)['state']
                    if (po.partner_type != 'external' and pol_state != 'draft') or (po.partner_type == 'external' and
                                                                                    pol_state not in ['draft', 'validated_n', 'validated']):
                        pol_sel_state = fields_obj.get_selection(cr, uid, 'purchase.order.line', 'state', pol_state, context=context)
                        partner_type = fields_obj.get_selection(cr, uid, 'res.partner', 'partner_type', po.partner_type, context=context)
                        if po.partner_type == 'external':
                            partner_msg = _('Draft, Validated-n and Validated states are authorized')
                        else:
                            partner_msg = _('Draft state is authorized')
                        line_err += _("The matching line can not be updated because its state is %s; only %s if the Partner is %s. ") \
                            % (pol_sel_state, partner_msg, partner_type)
                        pol_id = False
                else:
                    no_match_msg = [_('Line Number %s') % (line_num,)]
                    if prod_name:
                        no_match_msg.append(_("Product Code '%s'") % (prod_name,))
                    if comment:
                        no_match_msg.append(_("Comment '%s'") % (comment,))
                    if qty:
                        no_match_msg.append(_("Qty %s") % (qty,))
                    line_err += _('No matching PO line not already used by the import was found for: %s. ') \
                        % (', '.join(no_match_msg))

            if line_err:
                line_err = _('Line %s: ') % (row_num,) + line_err

            lines.append([pol_id, pol_data, line_err])

        wiz_state = 'done'
        imp_lines = 0
        for pol_id, line_data, line_err in lines:
            if not line_err:
                if pol_id:
                    pol_obj.write(cr, uid, pol_id, line_data, context=context)
                    imp_lines += 1
            else:
                lines_err.append(line_err)
        if lines_err:
            if lines_err:
                message = '%s:\n%s' % (_('Errors'), "\n".join(lines_err))
            wiz_state = 'error'

        end_time = time.time()
        total_time = str(round(end_time - start_time)) + _(' second(s)')
        final_message = _('''Importation completed in %s!
# of imported lines : %s on %s lines
# of ignored lines: %s
# of errors to correct: %s

%s''') % (total_time, imp_lines, len(lines), len(lines) - imp_lines, len(lines_err), message)
        self.write(cr, uid, wiz.id, {'state': wiz_state, 'message': final_message}, context=context)

        wb.close()  # Close manually because of readonly
        return True


wizard_update_po_line_import()
