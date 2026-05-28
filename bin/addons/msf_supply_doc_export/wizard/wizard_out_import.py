# -*- coding: utf-8 -*-
from osv import fields
from osv import osv
from tools.translate import _
from openpyxl import load_workbook
import base64
import json
from datetime import datetime
from io import BytesIO


class wizard_out_import(osv.osv_memory):
    _name = 'wizard.out.import'
    _rec_name = 'processor_id'

    _columns = {
        'processor_id': fields.many2one('outgoing.delivery.processor', string='OUT processor wizard', required=True, readonly=True),
        'file_to_import': fields.binary(string='File to import', filters='*.xls*'),
        'json_text': fields.text(string='JSON as text', help='Please put the data on a single line, with no line return'),
    }

    def normalize_data(self, cr, uid, data, json_import=False):
        if 'qty' in data:  # set to float
            if not data['qty']:
                data['qty'] = 0.0
            if isinstance(data['qty'], str):
                try:
                    data['qty'] = float(data['qty'])
                except:
                    if json_import:
                        raise osv.except_osv(_('Error'), _('Line %s: Key "product_qty" must be a number') % data['item'])
                    else:
                        raise osv.except_osv(_('Error'), _('Line %s: Column "Ordered Qty" must be a number') % data['item'])

        if 'kit' in data:  # set to str
            if not data['kit']:
                data['kit'] = ''

        if 'uom' in data:  # set to str
            if not data['uom']:
                data['uom'] = ''

        if 'batch' in data:  # set to str
            if not data['batch']:
                data['batch'] = ''

        if 'expiry_date' in data:
            if not data['expiry_date']:
                data['expiry_date'] = ''
            else:
                try:
                    if isinstance(data['expiry_date'], str):
                        data['expiry_date'] = datetime.strptime(data['expiry_date'], '%Y-%m-%d')
                    else:
                        data['expiry_date'] = datetime(data['expiry_date'].year, data['expiry_date'].month, data['expiry_date'].day)
                except:
                    if json_import:
                        raise osv.except_osv(_('Error'), _('Line %s: Key "expired_date" must be a date') % data['item'])
                    else:
                        raise osv.except_osv(_('Error'), _('Line %s: Column "Expiry Date" must be a date') % data['item'])

        return data

    def cancel(self, cr, uid, ids, context=None):
        if context is None:
            context = {}

        for wiz in self.read(cr, uid, ids, ['processor_id'], context=context):
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'outgoing.delivery.processor',
                'res_id': wiz['processor_id'],
                'view_type': 'form',
                'view_mode': 'form',
                'target': 'new',
            }

        return {'type': 'ir.actions.act_window_close'}

    def split_proc_move(self, cr, uid, proc_move_id, proc_move_qty, new_qty=0.00, context=None):
        """
        Split the line according to new parameters
        """
        if not proc_move_id:
            raise osv.except_osv(_('Error'), _('No line to split !'))

        proc_move_obj = self.pool.get('outgoing.delivery.move.processor')

        # New quantity must be greater than 0.00 and lower than the original move's qty
        if new_qty <= 0.00 or new_qty > proc_move_qty or new_qty == proc_move_qty:
            return False

        # Create a copy of the move with the new quantity
        context.update({'keepLineNumber': True})
        new_proc_move_id = proc_move_obj.copy(cr, uid, proc_move_id, {'ordered_quantity': new_qty, 'quantity': 0.00}, context=context)
        context.pop('keepLineNumber')

        # Update the original move
        update_qty = proc_move_qty - new_qty
        vals = {'ordered_quantity': update_qty}
        if context.get('sde_flow'):
            # Prevent adding unexpected qty to process
            vals['quantity'] = 0
        else:
            vals['quantity'] = proc_move_qty > update_qty and update_qty or proc_move_qty
        proc_move_obj.write(cr, uid, proc_move_id, vals, context=context)

        return new_proc_move_id

    def get_matching_proc_move(self, cr, uid, ids, line_data, product_id, wizard_id, treated_lines, context=None):
        if context is None:
            context = {}

        proc_move_obj = self.pool.get('outgoing.delivery.move.processor')

        proc_move_domain = [
            ('id', 'not in', treated_lines),
            ('wizard_id', '=', wizard_id),
            ('line_number', '=', line_data['item']),
            ('product_id', '=', product_id),
        ]
        exact_proc_move_domain = [x for x in proc_move_domain]
        exact_proc_move_domain.append(('ordered_quantity', '=', line_data['qty']))
        proc_move_ids = proc_move_obj.search(cr, uid, exact_proc_move_domain, limit=1, context=context)
        if proc_move_ids:
            exact_proc_move = proc_move_obj.browse(cr, uid, proc_move_ids[0], fields_to_fetch=['move_id'], context=context)
            if exact_proc_move.move_id.state != 'assigned':
                # Prevent modification of lines that are not Available
                return False
            else:
                return proc_move_ids[0]
        else:
            proc_move_ids = proc_move_obj.search(cr, uid, proc_move_domain, context=context)
            for proc_move in proc_move_obj.browse(cr, uid, proc_move_ids, fields_to_fetch=['ordered_quantity', 'move_id'], context=context):
                if 0 < line_data['qty'] < proc_move.ordered_quantity and proc_move.move_id.state == 'assigned':
                    return self.split_proc_move(cr, uid, proc_move.id, proc_move.ordered_quantity, line_data['qty'], context=context)
                else:
                    # Prevent modification of lines that are not Available
                    return False

        # To give a more specific error in case there is no match
        sql_treated_lines = """"""
        if treated_lines:
            if len(treated_lines) == 1:
                sql_treated_lines = """ AND mp.id != %s""" % (treated_lines[0],)
            else:
                sql_treated_lines = """ AND mp.id NOT IN %s""" % (tuple(treated_lines),)
        cr.execute("""
            SELECT mp.id
            FROM outgoing_delivery_move_processor mp LEFT JOIN stock_move m ON mp.move_id = m.id
            WHERE mp.wizard_id = %s AND m.state = 'assigned' AND mp.line_number = %s
        """ + sql_treated_lines, (wizard_id, line_data['item'])) # not_a_user_entry

        if not cr.fetchone():
            raise osv.except_osv(
                _('Error'), _('Line %s: Does not correspond to the ordered line. Please correct the Line Number of Product Code %s')
                % (line_data['item'], line_data['code'])
            )
        else:
            raise osv.except_osv(
                _('Error'), _('Line %s: Does not correspond to the ordered line. Please correct the Product Code %s')
                % (line_data['item'], line_data['code'])
            )

    def checks_on_batch(self, cr, uid, ids, product, line_data, context=None):
        if context is None:
            context = {}

        if product.batch_management and not line_data['batch']:
            raise osv.except_osv(
                _('Error'),
                _('Line %s: Product is batch number mandatory and no batch number is given') % line_data['item']
            )
        if not product.batch_management and product.perishable and not line_data['expiry_date']:
            raise osv.except_osv(
                _('Error'),
                _('Line %s: Product is expiry date mandatory and no expiry date is given') % line_data['item']
            )

    def get_data(self, cr, uid, ids, line_data, context=None):
        if context is None:
            context = {}

        prod_obj = self.pool.get('product.product')

        data = {}
        product_ids = prod_obj.search(cr, uid, [('default_code', '=ilike', line_data['code'])], limit=1, context=context)
        if not product_ids:
            raise osv.except_osv(_('Error'), _('Product with code %s not found in database') % (line_data['code'],))
        else:
            ftf = ['batch_management', 'perishable', 'type', 'subtype', 'default_code']
            product = prod_obj.browse(cr, uid, product_ids[0], fields_to_fetch=ftf, context=context)
            data.update({
                'product': product
            })

        return data

    def get_uom(self, cr, uid, ids, line_data, ordered_uom_category_id, json_import, context=None):
        if context is None:
            context = {}

        if json_import and not line_data.get('uom'):
            raise osv.except_osv(_('Error'), _('Line %s: Key "uom" is mandatory') % (line_data['item'],))

        uom_obj = self.pool.get('product.uom')
        uom_ids = uom_obj.search(cr, uid, [('name', '=ilike', line_data['uom']), ('category_id', '=', ordered_uom_category_id)],
                                 limit=1, context=context)
        if not uom_ids:
            raise osv.except_osv(_('Error'), _('UoM %s not found in database') % (line_data['uom'],))

        return uom_obj.browse(cr, uid, uom_ids[0], fields_to_fetch=['id', 'rounding'], context=context)

    def get_kit_id(self, cr, uid, ids, line_data, product_id, context=None):
        if context is None:
            context = {}

        kit_obj = self.pool.get('composition.kit')
        kit_domain = [('composition_reference', '=', line_data['kit']), ('composition_product_id', '=', product_id),
                      ('state', '=', 'completed'), ('composition_type', '=', 'real'), ('kcl_used_by', '=', False)]
        kit_ids = kit_obj.search(cr, uid, kit_domain, limit=1, context=context)
        if not kit_ids:
            raise osv.except_osv(_('Error'), _('No unused and completed Kit %s for the product %s found in database')
                                 % (line_data['kit'], line_data['code']))

        return kit_ids[0]

    def get_import_json_data(self, cr, uid, ids, json_data, context=None):
        if context is None:
            context = {}

        lines_data = {}
        line_index = 0
        for move in json_data.get('move_lines', []):
            line_index += 1
            lines_data[line_index] = {
                'item': move.get('line_number', False),
                'code': move.get('product_code', ''),
                'kit': move.get('kit', ''),
                'qty': move.get('product_qty', 0.0),
                'qty_to_process': move.get('qty_to_process', 0.0),
                'uom': move.get('uom', ''),
                'batch': move.get('prodlot_id', ''),
                'expiry_date': move.get('expired_date', ''),
            }

        return lines_data

    def get_import_data(self, cr, uid, ids, out_name, file_to_import, context=None):
        """
        Get the data from the XLSX file
        """
        if context is None:
            context = {}

        wb = load_workbook(filename=BytesIO(base64.b64decode(file_to_import)), read_only=True)
        sheet = wb.active

        # Get the reference
        out_ref = sheet[1] and sheet[1][1] and sheet[1][1].value or ''
        if out_ref.lower() != out_name.lower():
            wb.close()  # Close manually because of readonly
            raise osv.except_osv(_('Error'), _('OUT reference in the import file doesn\'t match with the current OUT'))

        # Fetch the data from the file
        lines_headers = ['item', 'code', 'description', 'comment', 'kit', 'src_location', 'dest_location',
                         'qty_in_stock', 'qty', 'qty_to_process', 'uom', 'batch', 'expiry_date', 'kc', 'dg', 'cs']
        lines_data = {}
        for cell in sheet.iter_rows(min_row=11, min_col=1, max_col=16):
            if not cell[0].value:  # Stop looking at lines if there is no data in first column
                break

            row_num = cell[0].row or 0
            line_data = {}
            for col, header in enumerate(lines_headers):
                line_data[header] = cell[col].value
            lines_data[row_num] = line_data

        wb.close()  # Close manually because of readonly

        return lines_data

    def import_out_xlsx(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        proc_move_obj = self.pool.get('outgoing.delivery.move.processor')
        prodlot_obj = self.pool.get('stock.production.lot')

        wiz = self.browse(cr, uid, ids[0], context=context)
        out = wiz.processor_id.picking_id
        json_import = False
        if wiz.json_text:
            json_import = True
            json_data = json.loads(wiz.json_text)
            lines_data = self.get_import_json_data(cr, uid, ids, json_data, context=context)
        else:
            if not wiz.file_to_import:
                raise osv.except_osv(_('Error'), _('Nothing to import.'))

            lines_data = self.get_import_data(cr, uid, ids, out.name, wiz.file_to_import, context=context)

        # Use the data that was fetched
        moves_data = []
        qty_per_line = {}
        treated_lines, ignored_lines = [], []
        for row_num, line_data in sorted(lines_data.items()):
            try:
                line_data['item'] = int(line_data['item'])
            except:
                if json_import:
                    raise osv.except_osv(_('Error'), _('Line %s: Key "line_number" must be an integer') % row_num)
                else:
                    raise osv.except_osv(_('Error'), _('File line %s: Column "Item" must be an integer') % row_num)

            # Completely ignore an imported line if qty_to_process is 'NULL' during JSON import
            if json_import and line_data.get('qty_to_process') == 'NULL':
                ignored_lines.append(_('%s (Qty to Process %s)') % (line_data['item'], line_data['qty_to_process']))
                continue
            if line_data['qty_to_process'] is None:
                if json_import:
                    raise osv.except_osv(_('Error'), _('Line %s: Key "qty_to_process" should contain the quantity to process and cannot be empty, please fill it with "0" instead') % line_data['item'])
                else:
                    raise osv.except_osv(_('Error'), _('Line %s: Column "Qty to Process" should contain the quantity to process and cannot be empty, please fill it with "0" instead') % line_data['item'])

            if not line_data['qty_to_process']:
                line_data['qty_to_process'] = 0.0
            if isinstance(line_data['qty_to_process'], str):
                try:
                    line_data['qty_to_process'] = float(line_data['qty_to_process'])
                except:
                    if json_import:
                        raise osv.except_osv(_('Error'), _('Line %s: Key "qty_to_process" must be a number') % line_data['item'])
                    else:
                        raise osv.except_osv(_('Error'), _('Line %s: Column "Qty to Process" must be a number') % line_data['item'])

            if line_data['qty_to_process'] and line_data['qty_to_process'] < 0:
                if json_import:
                    raise osv.except_osv(_('Error'), _('Line %s: Key "qty_to_process" should be greater than 0') % line_data['item'])
                else:
                    raise osv.except_osv(_('Error'), _('Line %s: Column "Qty to Process" should be greater than 0') % line_data['item'])

            if line_data['qty_to_process']:
                line_data = self.normalize_data(cr, uid, line_data, json_import=json_import)
                to_write = {}

                data = self.get_data(cr, uid, ids, line_data, context=context)
                product = data.get('product')
                proc_move_id = self.get_matching_proc_move(cr, uid, ids, line_data, product.id, wiz.processor_id.id,
                                                           treated_lines, context=context)
                if not proc_move_id:
                    ignored_lines.append(_('%s (Qty to Process %s)') % (line_data['item'], line_data['qty_to_process']))
                    continue
                else:
                    self.checks_on_batch(cr, uid, ids, product, line_data, context=context)
                    to_write.update({
                        'proc_move_id': proc_move_id,
                    })
                    if line_data['qty_to_process'] > line_data['qty']:
                        if json_import:
                            raise osv.except_osv(
                                _('Error'), _('Line %s: Key "qty_to_process" (%s) cannot be greater than "qty" (%s)')
                                % (line_data['item'], line_data['qty_to_process'], line_data['qty'])
                            )
                        else:
                            raise osv.except_osv(
                                _('Error'), _('Line %s: Column "Qty to Process" (%s) cannot be greater than "Ordered Qty" (%s)')
                                % (line_data['item'], line_data['qty_to_process'], line_data['qty'])
                            )
                    treated_lines.append(to_write['proc_move_id'])

                proc_move = proc_move_obj.browse(cr, uid, to_write['proc_move_id'], context=context)

                # Save qties by line
                if qty_per_line.get(line_data['item']):
                    qty_per_line[line_data['item']] += line_data['qty']
                else:
                    qty_per_line[line_data['item']] = line_data['qty']

                to_write['qty_to_process'] = line_data['qty_to_process']
                if proc_move.product_id.batch_management:
                    if line_data['batch'] and line_data['expiry_date']:
                        prodlot_ids = prodlot_obj.search(cr, uid, [('product_id', '=', proc_move.product_id.id),
                                                                   ('name', '=', line_data['batch']),
                                                                   ('life_date', '=', line_data['expiry_date'])], context=context)
                        if prodlot_ids:
                            to_write['prodlot_id'] = prodlot_ids[0]
                        else:
                            raise osv.except_osv(
                                _('Error'),
                                _('Line %s: The given batch number with this expiry date doesn\'t exist in database') % line_data['item']
                            )
                    else:
                        raise osv.except_osv(_('Error'),
                                             _('Line %s: Product %s must have a batch number and an expiry date')
                                             % (line_data['item'], line_data['code']))
                elif not proc_move.product_id.batch_management and proc_move.product_id.perishable and line_data['expiry_date']:
                    prodlot_ids = prodlot_obj.search(cr, uid, [('life_date', '=', line_data['expiry_date']),
                                                               ('type', '=', 'internal'),
                                                               ('product_id', '=', proc_move.product_id.id)], context=context)
                    if prodlot_ids:
                        to_write['prodlot_id'] = prodlot_ids[0]
                    else:
                        raise osv.except_osv(
                            _('Error'), _('Line %s: The given expiry date doesn\'t exist in database') % line_data['item']
                        )

                # UoM
                uom = self.get_uom(cr, uid, ids, line_data, proc_move.ordered_uom_category.id, json_import, context=context)
                if proc_move.ordered_uom_id.id != uom.id:
                    to_write['uom_id'] = uom.id
                if uom.rounding != 1:
                    to_write.update({'uom_rounding_is_pce': False, 'composition_list_id': False})
                if to_write.get('qty_to_process') and proc_move.ordered_uom_id.rounding != uom.rounding:
                    new_qty = self.pool.get('product.uom')._compute_round_up_qty(cr, uid, uom.id, to_write['qty_to_process'], context=context)
                    to_write['qty_to_process'] = new_qty

                # Kit
                if line_data['kit'] and proc_move.kit_check and product.type == 'product' and product.subtype == 'kit' \
                        and uom.rounding == 1:
                    to_write['composition_list_id'] = self.get_kit_id(cr, uid, ids, line_data, product.id, context=context)

                moves_data.append(to_write)

        cr.execute("""
            SELECT mp.line_number, p.default_code, SUM(mp.ordered_quantity) 
            FROM outgoing_delivery_move_processor mp, product_product p
            WHERE mp.product_id = p.id AND mp.wizard_id = %s
            GROUP BY mp.line_number, p.default_code
        """, (wiz.processor_id.id,))
        for prod in cr.fetchall():
            if context.get('sde_flow') and prod[2] != 0 and qty_per_line.get(prod[0]) and qty_per_line[prod[0]] > prod[2]:
                raise osv.except_osv(
                    _('Error'),
                    _('The total quantity of line #%s in the JSON data (%s) can not be more than the total qty on screen')
                    % (prod[0], qty_per_line.get(prod[0]))
                )
            elif not context.get('sde_flow') and prod[2] != 0 and qty_per_line.get(prod[0]) and qty_per_line[prod[0]] != prod[2]:
                raise osv.except_osv(
                    _('Error'),
                    _('The total quantity of line #%s in the import file (%s) doesn\'t match with the total qty on screen')
                    % (prod[0], qty_per_line.get(prod[0]))
                )

        for to_write in moves_data:
            if to_write.get('proc_move_id'):
                move_data = {'quantity': to_write.get('qty_to_process', 0)}
                if to_write.get('prodlot_id'):
                    move_data.update({'prodlot_id': to_write['prodlot_id']})
                if to_write.get('uom_id'):
                    move_data.update({'uom_id': to_write['uom_id']})
                if to_write.get('uom_rounding_is_pce'):
                    move_data.update({'uom_rounding_is_pce': to_write['uom_rounding_is_pce']})
                if to_write.get('composition_list_id'):
                    move_data.update({'composition_list_id': to_write['composition_list_id']})
                proc_move_obj.write(cr, uid, to_write['proc_move_id'], move_data, context=context)

        if context.get('sde_flow'):
            return ignored_lines
        else:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'outgoing.delivery.processor',
                'res_id': wiz.processor_id.id,
                'view_type': 'form',
                'view_mode': 'form',
                'target': 'new',
            }


wizard_out_import()
