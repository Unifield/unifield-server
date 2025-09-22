# -*- coding: utf-8 -*-
from osv import fields
from osv import osv
from tools.translate import _
from openpyxl import load_workbook
import base64
import tools
import time
from io import BytesIO
from datetime import datetime


class wizard_rfq_sent_import(osv.osv_memory):
    _name = 'wizard.rfq.sent.import'
    _rec_name = 'rfq_id'

    _columns = {
        'rfq_id': fields.many2one('purchase.order', string='Request for Quotation', required=True, readonly=True),
        'state': fields.selection([('draft', 'Draft'), ('in_progress', 'Import in progress'), ('error', 'Error'), ('done', 'Done')], string='State', readonly=True),
        'file_to_import': fields.binary(string='File to import', filters='*.xls*'),
        'filename': fields.char(size=64, string='Filename'),
        'message': fields.text(string='Message', readonly=True, translate=True),
    }

    def go_to_rfq(self, cr, uid, ids, context=None):
        '''
        Return to the initial view.
        '''
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        data_obj = self.pool.get('ir.model.data')
        for wiz_obj in self.read(cr, uid, ids, ['rfq_id'], context=context):
            rfq_id = wiz_obj['rfq_id']
            view_id = data_obj.get_object_reference(cr, uid, 'tender_flow', 'view_rfq_form')[1]
            tree_view_id = data_obj.get_object_reference(cr, uid, 'tender_flow', 'view_rfq_tree')[1]
            src_view_id = data_obj.get_object_reference(cr, uid, 'tender_flow', 'view_rfq_filter')[1]

            context.update({'rfq_ok': True, 'request_for_quotation': True})
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'purchase.order',
                'view_type': 'form',
                'view_mode': 'form,tree',
                'target': 'crush',
                'res_id': [rfq_id],
                'view_id': [view_id, tree_view_id],
                'search_view_id': src_view_id,
                'domain': [('rfq_ok', '=', True)],
                'context': context,
            }
        return True

    def dummy(self, cr, uid, ids, context=None):
        """
        This button is only for updating the view.
        """
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]
        for wiz in self.browse(cr, uid, ids, fields_to_fetch=['rfq_id', 'state'], context=context):
            if wiz.state not in ['done', 'error']:
                self.write(cr, uid, ids, {'message': _(' Import in progress... \n Please wait that the import is finished before editing %s.')
                                                     % (wiz.rfq_id.name,)})
        return False

    def import_file(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        if isinstance(ids, int):
            ids = [ids]

        pol_obj = self.pool.get('purchase.order.line')

        start_time = time.time()
        wiz = self.browse(cr, uid, ids[0], context=context)
        rfq = wiz.rfq_id
        if rfq.rfq_state != 'sent':
            raise osv.except_osv(_('Error'), _('This import can only be used on Sent RfQs'))
        if not wiz.file_to_import:
            raise osv.except_osv(_('Error'), _('Nothing to import.'))
        self.write(cr, uid, wiz.id, {'state': 'in_progress'}, context=context)

        wb = load_workbook(filename=BytesIO(base64.b64decode(wiz.file_to_import)), read_only=True)
        sheet = wb.active

        # Check the Reference
        if not sheet['B1'].value or sheet['B1'].value != rfq.name:
            end_time = time.time()
            total_time = str(round(end_time - start_time)) + _(' second(s)')
            final_message = _('''Importation completed in %s!
# of errors to correct: 1

The Order Reference in the file (%s) does not match the one in the RFQ (%s)''') % (total_time, sheet['B1'].value or _('None'), rfq.name)
            self.write(cr, uid, wiz.id, {'state': 'error', 'message': final_message}, context=context)

            wb.close()  # Close manually because of readonly
            return True

        # Valid Till
        header_write = {}
        header_err = ''
        vt_date = sheet['B4'].value
        if vt_date:
            if sheet['B4'].data_type != 'd':
                header_err = _('Valid Till must be a date. ')
            else:
                header_write['valid_till'] = vt_date.strftime('%Y-%m-%d')
        else:
            header_write['valid_till'] = False

        lines = []
        lines_err = []
        message = ''
        used_rfql = []

        for cell in sheet.iter_rows(min_row=13, min_col=0, max_col=10):
            line_num = cell[0].value
            if not line_num:  # Stop looking at lines if there is no line_number
                break

            line_err = ''
            row_num = cell[0].row or ''

            # Line Number
            if cell[0].data_type != 'n':
                try:
                    line_num = int(line_num)
                except ValueError:
                    line_num = False
                    line_err += _('The Line Number must be a number. ')

            if line_num and line_num in used_rfql:
                line_num = False
                line_err += _('The same Line Number has already been used. ')

            rfql_data = {}
            rfql_id = False

            if line_num:
                # Unit Price
                price = cell[5].value
                if price:
                    if cell[5].data_type != 'n':
                        try:
                            price = float(price.rstrip().replace(',', '.'))
                        except ValueError:
                            line_err += _('The Unit Price must be a number. ')
                    if isinstance(price, float) or isinstance(price, int):
                        if price < 0:
                            line_err += _('The Unit Price must be positive. ')
                        else:
                            rfql_data.update({'price_unit': price})
                else:
                    line_err += _('The Unit Price is mandatory for each line. ')

                # Confirmed Delivery Date
                cdd = cell[7].value
                if cdd:
                    if cell[7].data_type != 'd':
                        line_err += _('The Confirmed Delivery Date must be a date. ')
                    else:
                        rfql_data.update({'confirmed_delivery_date': cdd.strftime('%Y-%m-%d')})
                else:
                    rfql_data.update({'confirmed_delivery_date': False})

                if line_err:
                    rfql_id = False
                else:
                    rfql_domain = [('order_id', '=', rfq.id), ('line_number', '=', line_num),
                                   ('rfq_line_state', 'not in', ['cancel', 'cancel_r'])]
                    rfql_ids = pol_obj.search(cr, uid, rfql_domain, context=context)
                    if rfql_ids:
                        used_rfql.append(line_num)
                        rfql_id = rfql_ids[0]
                    else:
                        line_err += _('No RfQ line was found for the Line Number %s. ') % (line_num,)

            if line_err:
                line_err = _('Line %s: ') % (row_num,) + line_err

            lines.append([rfql_id, rfql_data, line_err])

        wiz_state = 'done'
        imp_lines = 0
        if header_write:
            self.pool.get('purchase.order').write(cr, uid, rfq.id, header_write, context=context)
        for rfql_id, line_data, line_err in lines:
            if not line_err:
                if rfql_id:
                    pol_obj.write(cr, uid, rfql_id, line_data, context=context)
                    imp_lines += 1
            else:
                lines_err.append(line_err)
        if lines_err or header_err:
            if header_err:
                lines_err.insert(0, header_err)
            message = '%s:\n%s' % (_('Errors'), "\n".join(lines_err))
            wiz_state = 'error'

        end_time = time.time()
        total_time = str(round(end_time - start_time)) + _(' second(s)')
        final_message = _('''Importation completed in %s!
# of imported lines : %s on %s lines
# of ignored lines: %s
# of errors to correct: %s

%s''') % (total_time, imp_lines, len(lines), len(lines) - imp_lines, len(lines_err), message)
        self.write(cr, uid, wiz.id, {'state': wiz_state, 'message': final_message}, context=context)

        wb.close()  # Close manually because of readonly
        return True


wizard_rfq_sent_import()
