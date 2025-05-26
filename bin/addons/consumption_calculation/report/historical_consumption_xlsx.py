# -*- coding: utf-8 -*-
from spreadsheet_xml.xlsx_write import XlsxReport
from spreadsheet_xml.xlsx_write import XlsxReportParser
from datetime import datetime
from tools.translate import _
from openpyxl.utils.cell import get_column_letter


class historical_parser(XlsxReportParser):

    def generate(self, context=None):
        if context is None:
            context = {}

        context['limit_location'] = 0
        history_obj = self.pool.get('product.history.consumption')
        product_obj = self.pool.get('product.product')
        month_fields_to_read = []
        h_amc = history_obj.browse(self.cr, self.uid, self.ids[0], context=context)

        prod_dom, history_ctx = history_obj.get_data(self.cr, self.uid, [self.ids[0]], context=context)
        list_months = history_obj.get_months(self.cr, self.uid, [self.ids[0]], context=context)

        sheet = self.workbook.active

        self.create_style_from_template('title_style', 'A1')
        self.create_style_from_template('header_style', 'C2')
        self.create_style_from_template('header_date_style', 'D2')

        self.create_style_from_template('sub_header_style', 'A3')

        self.create_style_from_template('prod_style', 'A4')
        self.create_style_from_template('amc_style', 'C4')

        idx = 1
        if h_amc.remove_negative_amc:
            sheet.row_dimensions[1].height = 24
            idx = 2
            row_index = 3
        else:
            row_index = 2

        if h_amc.consumption_type == 'rr-amc':
            if h_amc.txt_source and len(h_amc.txt_source) > 30:
                sheet.row_dimensions[idx].height = 50
            else:
                sheet.row_dimensions[idx].height = 20

            if h_amc.txt_destination and len(h_amc.txt_destination) > 30:
                sheet.row_dimensions[idx+1].height = 50
            else:
                sheet.row_dimensions[idx+1].height = 20

            if h_amc.txt_ext_partner and len(h_amc.txt_ext_partner) > 30:
                sheet.row_dimensions[idx+1].height = 50
            else:
                sheet.row_dimensions[idx+1].height = 20
            row_index += 3

        self.duplicate_column_dimensions(default_width=10.75)
        sheet.freeze_panes = 'D%d' % (row_index+1)

        col = 3
        col_width = self.workbook_template.active.column_dimensions['C'].width
        for month in list_months:
            col += 1
            sheet.column_dimensions[get_column_letter(col)].width = col_width

        sheet.title = _('Historical Consumption')
        if h_amc.remove_negative_amc:
            sheet.append([self.cell_ro(_('This report hides negative AMC / MCs (they are set to 0)'), 'title_style')])
            sheet.merged_cells.ranges.append("A1:C1")

        if h_amc.consumption_type == 'rr-amc':
            sheet.append([
                self.cell_ro(_('Source Locations'), 'sub_header_style'),
                self.cell_ro(h_amc.txt_source, 'sub_header_style'),
                self.cell_ro('', 'sub_header_style'),
            ])
            sheet.merged_cells.ranges.append("B%(idx)s:C%(idx)s" % {'idx': row_index-4})
            sheet.append([
                self.cell_ro(_('Destination Locations'), 'sub_header_style'),
                self.cell_ro(h_amc.txt_destination, 'sub_header_style'),
                self.cell_ro('', 'sub_header_style'),
            ])
            sheet.merged_cells.ranges.append("B%(idx)s:C%(idx)s" % {'idx': row_index-3})
            sheet.append([
                self.cell_ro(_('External Partners'), 'sub_header_style'),
                self.cell_ro(h_amc.txt_ext_partner, 'sub_header_style'),
                self.cell_ro('', 'sub_header_style'),
            ])
            sheet.merged_cells.ranges.append("B%(idx)s:C%(idx)s" % {'idx': row_index-2})


        nb_month = len(list_months) + 3
        sheet.auto_filter.ref = "A%(idx)d:%(last_month)s%(idx)d" % {'idx': row_index, 'last_month': get_column_letter(nb_month)}

        row_index += 1

        header = ['', '', self.cell_ro(_('Entire Period'), 'header_style')]
        for month in list_months:
            dt_month = datetime.strptime(month.get('date_from'), '%Y-%m-%d')
            header.append(self.cell_ro(dt_month, 'header_date_style'))
            month_fields_to_read.append(dt_month.strftime('%m_%Y'))

        sheet.append(header)

        if h_amc.consumption_type == 'amc':
            full_title = _('AMC')
            sub_title = _('MC')
        elif h_amc.consumption_type == 'rr-amc':
            if h_amc.adjusted_rr_amc:
                full_title = _('Adj. RR-AMC')
                sub_title = _('Adj. RR-MC')
            else:
                full_title = _('RR-AMC')
                sub_title = _('RR-MC')
        else:
            full_title = _('RAC')
            sub_title = _('RC')

        sub_header = [
            self.cell_ro(_('CODE'), 'sub_header_style'),
            self.cell_ro(_('DESCRIPTION'), 'sub_header_style'),
            self.cell_ro(full_title, 'sub_header_style'),
        ]
        for month in list_months:
            sub_header.append(self.cell_ro(sub_title, 'sub_header_style'))
        sheet.append(sub_header)

        if not h_amc.remove_negative_amc:
            prod_dom.append(('average', '!=', 0))
        else:
            prod_dom.append(('average', '>', 0))
        prod_ids = product_obj.search(self.cr, self.uid, prod_dom, context=history_ctx)
        max_read = 500
        prod_ctx = context.copy()
        prod_ctx.update(history_ctx)
        offset = 0

        row_height = self.workbook_template.active.row_dimensions[4].height
        while offset <= len(prod_ids):
            for prod in product_obj.read(self.cr, self.uid, prod_ids[offset:max_read+offset], ['default_code', 'name', 'average'], context=prod_ctx):
                prod_data = [self.cell_ro(prod['default_code'], 'prod_style'), self.cell_ro(prod['name'], 'prod_style'), self.cell_ro(prod.get('average', 0), 'amc_style')]
                for m in month_fields_to_read:
                    prod_data.append(self.cell_ro(prod.get(m, 0), 'amc_style'))
                sheet.row_dimensions[row_index].height = row_height
                sheet.append(prod_data)
                row_index += 1
            offset+=max_read


XlsxReport('report.report_historical_consumption_xlsx', parser=historical_parser, template='addons/consumption_calculation/report/historical_consumption.xlsx')

