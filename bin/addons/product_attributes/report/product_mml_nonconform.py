# -*- coding: utf-8 -*-
from spreadsheet_xml.xlsx_write import XlsxReport
from spreadsheet_xml.xlsx_write import XlsxReportParser
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from tools.translate import _


class common_non_conform(XlsxReportParser):
    def generate(self, context=None):
        if context is None:
            context = {}

        data_obj = self.pool.get('ir.model.data')
        location_obj = self.pool.get('stock.location')

        company_record = self.pool.get('res.company')._get_instance_record(self.cr, self.uid)
        instance_code = company_record.code
        self.instance_id = company_record.id

        sheet = self.workbook.active

        self.create_style_from_template('title_style', 'A1')
        self.create_style_from_template('filter_txt', 'A2')
        self.create_style_from_template('filter_date', 'B3')


        self.create_style_from_template('header_row', 'A5')
        self.create_style_from_template('stock_row', 'H5')
        self.create_style_from_template('date_row', 'X6')
        self.create_style_from_template('row', 'A6')


        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 27
        sheet.column_dimensions['C'].width = 11
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10

        #self.duplicate_column_dimensions(default_width=10.75)
        sheet.freeze_panes = 'C6'


        title = self.get_title()
        sheet.title = _(title)

        sheet.merged_cells.ranges.append("A1:F1")
        sheet.append([self.cell_ro(_(title), 'title_style')])

        sheet.append([
            self.cell_ro(_('Instance'), 'filter_txt'),
            self.cell_ro(instance_code, 'filter_txt'),
        ])
        sheet.append([
            self.cell_ro(_('Date'), 'filter_txt'),
            self.cell_ro(datetime.now(), 'filter_date'),
        ])
        sheet.append([])
        extra_col = self.extra_col()

        headers = [
            self.cell_ro(_('Code'), 'header_row'),
            self.cell_ro(_('Description'), 'header_row'),
            self.cell_ro(_('Product Creator'), 'header_row'),
            self.cell_ro(_('Standardization Level'), 'header_row'),
            self.cell_ro(_('UniData Status'), 'header_row'),
            self.cell_ro(_('UniField Status'), 'header_row'),
            self.cell_ro(_(extra_col['label']), 'header_row'),
            self.cell_ro(_('Instance stock'), 'stock_row'),
            self.cell_ro(_('Instance stock val.'), 'stock_row'),
            self.cell_ro(_('Stock Qty.'), 'stock_row')
        ]

        stock_location_id = data_obj.get_object_reference(self.cr, self.uid, 'stock', 'stock_location_stock')[1]
        cross_loc = location_obj.search(self.cr, self.uid, [('cross_docking_location_ok', '=', True)], context=context)
        internal_loc = location_obj.search(self.cr, self.uid, [('usage', '=', 'internal')], context=context)
        stock_loc = location_obj.search(self.cr, self.uid, [('location_id', 'child_of', stock_location_id),
                                                            ('id', 'not in', cross_loc),
                                                            ('central_location_ok', '=', False)], context=context)
        location_ids = location_obj.search(self.cr, self.uid, [('usage', '=', 'internal')], context=context)


        col_index = 10
        quarantine_ids = []
        displayed_location_ids = []
        for x in location_obj.read(self.cr, self.uid, location_ids, ['name', 'quarantine_location'], context=context):
            sheet.column_dimensions[get_column_letter(col_index)].width = 11
            if x['quarantine_location']:
                quarantine_ids.append(x['id'])
                if len(quarantine_ids) > 1:
                    continue
                displayed_location_ids.append('Q')
                x['name'] = _('Quarantine / For Scrap Qty')
            else:
                displayed_location_ids.append(x['id'])
            col_index += 1
            headers.append(self.cell_ro(x['name'], 'stock_row'))

        sheet.column_dimensions[get_column_letter(col_index)].width = 10
        sheet.column_dimensions[get_column_letter(col_index+1)].width = 10
        headers += [
            self.cell_ro(_('In Pipe Qty'), 'stock_row'),
            self.cell_ro(_('Date of last product transaction'), 'header_row'),
        ]
        sheet.append(headers)

        prod_obj = self.pool.get('product.product')

        fields = prod_obj.fields_get(self.cr, self.uid, ['standard_ok', 'state_ud', extra_col['field']], context=context)
        label_standard_ok = dict(fields['standard_ok']['selection'])
        label_state_ud = dict(fields['state_ud']['selection'])
        label_extra_col = dict(fields[extra_col['field']]['selection'])

        page_size = 250
        offset = 0
        progress = 0

        bk_id = self.context.get('background_id')

        query = self.get_query()

        while True:
            self.cr.execute(query+ "offset %s limit %s", (offset, page_size))
            p_ids = [x[0] for x in self.cr.fetchall()]
            if not p_ids:
                break
            progress += 5
            if bk_id:
                self.pool.get('memory.background.report').write(self.cr, self.uid, bk_id, {'percent': min(progress, 95)/100.})

            offset += len(p_ids)

            loc_detail = {}
            self.cr.execute('''
                select product_id, location_id, quantity
            from stock_mission_report_line_location
            where
                product_id in %s
                and location_id in %s
            ''', (tuple(p_ids), tuple(location_ids)))
            for x in self.cr.fetchall():
                loc_detail.setdefault(x[0], {}).setdefault(x[1], x[2])

            last_move = {}
            self.cr.execute('''
                select
                    max(m.date), m.product_id
                from
                    stock_move m
                left join stock_picking p on p.id = m.picking_id
                where
                    coalesce(p.type, 'in') = 'in'
                    and m.product_id in %s
                    and m.state ='done'
                group by product_id
            ''', (tuple(p_ids,), ))

            for x in self.cr.fetchall():
                if x[0]:
                    last_move[x[1]] = datetime.strptime(x[0], '%Y-%m-%d %H:%M:%S')

            self.cr.execute('''
                select
                    max(po.create_date), po.product_id
                from
                    purchase_order_line po
                where
                    po.product_id in %s
                    and po.state not in ('cancel', 'cancel_r')
                group by po.product_id
            ''', (tuple(p_ids, ), ))

            for x in self.cr.fetchall():
                if x[0]:
                    st_date = datetime.strptime(x[0].split('.')[0], '%Y-%m-%d %H:%M:%S')
                    if not last_move.get(x[1]) or st_date > last_move.get(x[1]):
                        last_move[x[1]] = st_date

            progress += 5
            if bk_id:
                self.pool.get('memory.background.report').write(self.cr, self.uid, bk_id, {'percent': min(progress, 95)/100.})

            # ctx change to include Cross Dock in PO pipe qty
            ctx = context.copy()
            ctx['location_category'] = ['stock', 'consumption_unit', 'eprep', 'transition']
            ctx['location_usage'] = ['internal']
            for prod in prod_obj.browse(self.cr, self.uid, p_ids, fields_to_fetch=['default_code', 'name', 'international_status', 'standard_ok', 'state_ud', 'state', extra_col['field'], 'standard_price', 'incoming_qty'], context=ctx):
                line = [
                    self.cell_ro(prod.default_code, 'row'),
                    self.cell_ro(prod.name, 'row'),
                    self.cell_ro(prod.international_status.name, 'row'),
                    self.cell_ro(label_standard_ok.get(prod.standard_ok), 'row'),
                    self.cell_ro(label_state_ud.get(prod.state_ud), 'row'),
                    self.cell_ro(prod.state.name, 'row'),
                    self.cell_ro(label_extra_col.get(prod[extra_col['field']]), 'row'),
                ]
                stock_qty = 0
                for x in stock_loc:
                    stock_qty += loc_detail.get(prod.id, {}).get(x) or 0
                internal_qty = 0
                for x in internal_loc:
                    internal_qty += loc_detail.get(prod.id, {}).get(x) or 0
                loc_detail.setdefault(prod.id, {}).setdefault('Q', 0)
                for q_id in quarantine_ids:
                    loc_detail[prod.id]['Q'] += loc_detail.get(prod.id, {}).get(q_id) or 0
                line += [
                    self.cell_ro(internal_qty or '', 'row'),
                    self.cell_ro(round(internal_qty * prod.standard_price, 2) or '', 'row'),
                    self.cell_ro(stock_qty or '', 'row'),
                ]
                for loc_id in displayed_location_ids:
                    line.append(self.cell_ro(loc_detail.get(prod.id, {}).get(loc_id) or '', 'row'))
                line.append(self.cell_ro(prod.incoming_qty or '', 'row'))
                line.append(self.cell_ro(last_move.get(prod.id) or '', 'date_row'))
                sheet.append(line)

            if len(p_ids) < page_size:
                break

        sheet.auto_filter.ref = "A5:G5"

class product_mml_nonconform(common_non_conform):
    def get_query(self):
        extra_join = ''
        extra_cond = ''
        if self.pool.get('non.conform.inpipe').read(self.cr, self.uid, self.ids[0], ['include_pipe'])['include_pipe']:
            extra_join = '''
            left join purchase_order_line pol on pol.product_id = p.id and pol.state in ('validated', 'validated_n', 'sourced_sy', 'sourced_v', 'sourced_n')
            left join (
                select m.product_id as product_id
                from stock_move m, stock_location l2
                where
                l2.id = m.location_dest_id
                and l2.usage = 'internal'
                and m.state in ('confirmed' ,'assigned')
                and m.product_qty > 0
            ) inc on inc.product_id = p.id
            '''
            extra_cond = 'or count(pol.id) > 0 or count(inc.product_id) > 0'
        return """
            select p.id
            from product_product p
            left join product_template tmpl on tmpl.id = p.product_tmpl_id
            left join product_nomenclature nom on tmpl.nomen_manda_0 = nom.id
            left join (
                select l.product_id as product_id
                from
                stock_mission_report_line_location l, stock_location location
                where
                    l.quantity > 0
                    and location.id = l.location_id
                    and  location.usage = 'internal'
            ) smr on smr.product_id = p.id
            left join product_project_rel p_rel on p.id = p_rel.product_id
            left join product_country_rel c_rel on p_rel is null and c_rel.product_id = p.id
            left join unidata_project up1 on up1.id = p_rel.unidata_project_id or up1.country_id = c_rel.unidata_country_id
            """ +  extra_join  + """
            where
                nom.name='MED'
                and nom.level = 0
            group by p.id
            HAVING
                (
                    (count(smr.product_id) > 0 """ + extra_cond + """)
                    and (
                        bool_and(coalesce(oc_validation,'f'))='f'
                        or
                        not ARRAY[%s]<@array_agg(up1.instance_id)
                        and
                        count(up1.instance_id)>0
                    )
                 )
            order by p.default_code
        """ % self.instance_id

    def get_title(self):
        return _('Products not in MML')

    def extra_col(self):
        return {'label': _('In MSL?'), 'field': 'msl_status'}

class product_msl_nonconform(common_non_conform):
    def get_query(self):
        ud_proj = self.pool.get('unidata.project').search(self.cr, self.uid, [('uf_active', '=', 't'), ('instance_id', '=', self.instance_id)])
        if not ud_proj:
            ud_proj = [0]
        extra_join = ''
        extra_cond = ''
        if self.pool.get('non.conform.inpipe').read(self.cr, self.uid, self.ids[0], ['include_pipe'])['include_pipe']:
            extra_join = '''
            left join purchase_order_line pol on pol.product_id = p.id and pol.state in ('validated', 'validated_n', 'sourced_sy', 'sourced_v', 'sourced_n')
            left join (
                select m.product_id as product_id
                from stock_move m, stock_location l2
                where
                l2.id = m.location_dest_id
                and l2.usage = 'internal'
                and m.state in ('confirmed' ,'assigned')
                and m.product_qty > 0
            ) inc on inc.product_id = p.id
            '''

            extra_cond = 'or count(pol.id) > 0 or count(inc.product_id) > 0'

        return """
            select p.id
            from product_product p
            left join product_template tmpl on tmpl.id = p.product_tmpl_id
            left join product_nomenclature nom on tmpl.nomen_manda_0 = nom.id
            left join (
                select l.product_id as product_id
                from
                stock_mission_report_line_location l, stock_location location
                where
                    l.quantity > 0
                    and location.id = l.location_id
                    and  location.usage = 'internal'
            ) smr on smr.product_id = p.id
            """ + extra_join + """
            left join product_msl_rel msl_rel on msl_rel.product_id = p.id and msl_rel.creation_date is not null and msl_rel.msl_id = """ + str(ud_proj[0]) + """
            left join product_international_status creator on creator.id = p.international_status
            where
                nom.name='MED'
                and nom.level = 0
                and msl_rel.product_id is null
                and creator.code = 'unidata'
            group by p.id
            HAVING
                (
                    count(smr.product_id) > 0 """ + extra_cond + """
                )
            order by p.default_code
        """

    def get_title(self):
        return _('Products not in MSL')

    def extra_col(self):
        return {'label': _('In MML?'), 'field': 'mml_status'}


XlsxReport('report.report.product_mml_nonconform', parser=product_mml_nonconform, template='addons/product_attributes/report/product_mml_nonconform.xlsx')
XlsxReport('report.report.product_msl_nonconform', parser=product_msl_nonconform, template='addons/product_attributes/report/product_mml_nonconform.xlsx')

